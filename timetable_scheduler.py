"""
timetable_scheduler.py
======================
Reads Planning for Timetable.xlsx and writes Timetable_Output.xlsx with all
classes auto-scheduled.

Usage:
    python timetable_scheduler.py

Each new Term: update the input Excel data, then re-run the script.
"""

import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

# ─── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE   = os.path.join(SCRIPT_DIR, "Planning for Timetable.xlsx")
OUTPUT_FILE  = os.path.join(SCRIPT_DIR, "Timetable_Output.xlsx")

# ─── Constants ────────────────────────────────────────────────────────────────

PREFERRED_DAYS = ["Monday", "Tuesday", "Thursday"]
ALL_DAYS       = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Valid slot pairs for 4-hr subjects, in preference order
SLOT_PAIRS_4HR = [
    ("0900 - 1100", "1100 - 1300"),   # morning block (best)
    ("1400 - 1600", "1600 - 1800"),   # afternoon block
    ("0900 - 1100", "1400 - 1600"),   # split morning + early afternoon
    ("1100 - 1300", "1400 - 1600"),   # split late morning + afternoon
    ("0900 - 1100", "1600 - 1800"),   # split morning + late afternoon
    ("1100 - 1300", "1600 - 1800"),   # split late morning + late afternoon
]

# Single slots for 2-hr subjects (Mathematics)
SLOTS_2HR = ["0900 - 1100", "1100 - 1300", "1400 - 1600", "1600 - 1800"]

# Exact sheet names (note trailing space on Monday)
DAY_TO_SHEET = {
    "Monday":    "Mon(Term1) ",
    "Tuesday":   "Tue(Term1)",
    "Wednesday": "Wed(Term1)",
    "Thursday":  "Thu(Term1)",
    "Friday":    "Fri(Term1)",
}

# Starting column (1-based) of each time slot group in daily timetable sheets
# Each group occupies 5 columns: Code, ClassSize, Subject, Lecturer, Time
SLOT_TO_START_COL = {
    "0900 - 1100": 5,    # col E
    "1100 - 1300": 10,   # col J
    "1400 - 1600": 22,   # col V
    "1600 - 1800": 27,   # col AA
}

# Map class centre suffix -> room code prefix for preferred room matching
CLASS_CENTRE_TO_ROOM_PREFIX = {
    "CS":  "CSW",
    "WT":  "WT",
    "SW":  "SW",
    "TW":  "TW",
    "TM":  "TM",
    "TS":  "TS",
    "KT":  "KT",
    "TK":  "TKO",
    "ST":  "ST",
    "FL":  "FL",
}

# ─── Data classes ─────────────────────────────────────────────────────────────

@dataclass
class ClassEntry:
    code:         str   # "DAE101_CS1"
    subject_code: str   # "DAE101"
    subject_name: str   # "Chinese Language"
    subject_cn:   str   # "中國語文"
    loading:      int   # 2 or 4 (hours)
    students:     int
    centre:       str   # "CS" extracted from code

@dataclass
class TeacherEntry:
    name:           str
    subject_quotas: dict = field(default_factory=dict)  # {subject_name: max_classes}

@dataclass
class RoomEntry:
    code:     str   # "CSW - C1"
    capacity: int
    prefix:   str   # "CSW" (part before " - ")

@dataclass
class ScheduledClass:
    entry:     ClassEntry
    day:       str
    slot1:     str
    slot2:     Optional[str]   # None for 2-hr subjects
    room:      str
    lecturer1: str
    lecturer2: Optional[str] = None
    lecturer3: Optional[str] = None


# ─── Phase 1 — Read inputs ────────────────────────────────────────────────────

def _extract_centre(class_code: str) -> str:
    """'DAE101_CS1' -> 'CS',  'DAE102_WT2' -> 'WT'"""
    parts = class_code.split("_")
    if len(parts) >= 2:
        return re.sub(r"\d+$", "", parts[1])
    return ""


def read_classes(wb) -> list:
    ws = wb["Class list"]
    classes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        subject_code = str(code).split("_")[0]
        subject_name = str(row[1]).strip() if row[1] else ""
        subject_cn   = str(row[2]).strip() if row[2] else ""
        loading      = int(row[6]) if isinstance(row[6], (int, float)) else 4
        students     = int(row[7]) if isinstance(row[7], (int, float)) else 0
        centre       = _extract_centre(str(code))
        classes.append(ClassEntry(
            code=str(code), subject_code=subject_code,
            subject_name=subject_name, subject_cn=subject_cn,
            loading=loading, students=students, centre=centre,
        ))
    return classes


_SUBJECT_NAME_MAP = {
    "MathsPlus":              "Maths Plus",
    "Career and Life Learning": "Career and Life Planning",
}

def read_teachers(wb) -> list:
    ws = wb["Teacher load table with subject"]
    raw_headers  = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Normalise subject column names to match Class list sheet
    headers      = [_SUBJECT_NAME_MAP.get(h, h) for h in raw_headers]
    subject_cols = headers[2:]  # columns C onward are subjects

    teachers: dict = {}  # name -> TeacherEntry
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        if name not in teachers:
            teachers[name] = TeacherEntry(name=name)
        for i, subject in enumerate(subject_cols):
            if not subject:
                continue
            val = row[i + 2] if (i + 2) < len(row) else None
            if isinstance(val, (int, float)) and val > 0:
                capped   = min(int(val), 20)   # cap data errors like "60"
                existing = teachers[name].subject_quotas.get(subject, 0)
                teachers[name].subject_quotas[subject] = existing + capped
    return list(teachers.values())


def read_rooms(wb) -> list:
    ws = wb["Centre Room Allocation"]
    rooms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code     = row[0]
        capacity = row[2]   # "No of Seats" column
        if not code or not isinstance(capacity, (int, float)) or capacity <= 0:
            continue
        code_str = str(code).strip()
        prefix   = code_str.split(" - ")[0].strip() if " - " in code_str else code_str
        rooms.append(RoomEntry(code=code_str, capacity=int(capacity), prefix=prefix))
    return rooms


# ─── Phase 2 — Assign teachers ────────────────────────────────────────────────

def assign_teachers(classes: list, teachers: list) -> tuple:
    """
    Round-robin assignment of Lecturer1 for each class, respecting quotas.
    Returns (assignments dict {code: teacher_name}, unassigned list).
    """
    # Build subject -> eligible teachers, sorted by quota desc
    subject_teachers: dict = defaultdict(list)
    for t in teachers:
        for subj, quota in t.subject_quotas.items():
            if quota > 0:
                subject_teachers[subj].append(t)
    for subj in subject_teachers:
        subject_teachers[subj].sort(
            key=lambda t: t.subject_quotas.get(subj, 0), reverse=True
        )

    remaining: dict = {}
    for t in teachers:
        for subj, quota in t.subject_quotas.items():
            remaining[(t.name, subj)] = quota

    by_subject: dict = defaultdict(list)
    for c in classes:
        by_subject[c.subject_name].append(c)

    assignments: dict = {}
    unassigned:  list = []

    for subj, subj_classes in by_subject.items():
        eligible = subject_teachers.get(subj, [])
        if not eligible:
            unassigned.extend(c.code for c in subj_classes)
            continue

        teacher_idx = 0
        for c in subj_classes:
            assigned = False
            for offset in range(len(eligible)):
                t = eligible[(teacher_idx + offset) % len(eligible)]
                if remaining.get((t.name, subj), 0) > 0:
                    assignments[c.code] = t.name
                    remaining[(t.name, subj)] -= 1
                    teacher_idx = (teacher_idx + offset + 1) % len(eligible)
                    assigned = True
                    break
            if not assigned:
                # All quotas exhausted — overflow to first teacher
                t = eligible[teacher_idx % len(eligible)]
                assignments[c.code] = t.name
                teacher_idx = (teacher_idx + 1) % len(eligible)

    return assignments, unassigned


# ─── Phase 3 — Greedy scheduler ───────────────────────────────────────────────

def _preferred_days(c: ClassEntry) -> list:
    """CS1-3 (Police Cadet) may use any day; everything else prefers Mon/Tue/Thu."""
    is_police_cadet = (c.centre == "CS" and
                       any(c.code.endswith(x) for x in ("CS1", "CS2", "CS3")))
    if is_police_cadet:
        return list(ALL_DAYS)
    return PREFERRED_DAYS + [d for d in ALL_DAYS if d not in PREFERRED_DAYS]


def _rooms_by_fit(rooms: list, students: int, preferred_prefix: str) -> list:
    """Fitting rooms: preferred centre first, then smallest-that-fits order."""
    fitting = [r for r in rooms if r.capacity >= students]
    preferred = sorted([r for r in fitting if r.prefix == preferred_prefix],
                       key=lambda r: r.capacity)
    others    = sorted([r for r in fitting if r.prefix != preferred_prefix],
                       key=lambda r: r.capacity)
    return preferred + others


def schedule_classes(classes: list, teacher_assignments: dict, rooms: list) -> tuple:
    teacher_busy: dict = defaultdict(lambda: defaultdict(set))
    room_busy:    dict = defaultdict(lambda: defaultdict(set))
    centre_day:   dict = defaultdict(lambda: defaultdict(int))

    # Largest classes first, 4-hr before 2-hr
    sorted_classes = sorted(classes, key=lambda c: (-c.students, -c.loading))

    scheduled:   list = []
    unscheduled: list = []

    for c in sorted_classes:
        teacher = teacher_assignments.get(c.code)
        if not teacher:
            unscheduled.append(c.code)
            continue

        preferred_prefix = CLASS_CENTRE_TO_ROOM_PREFIX.get(c.centre, c.centre)
        slot_options = (SLOT_PAIRS_4HR if c.loading == 4
                        else [(s, None) for s in SLOTS_2HR])

        # Sort preferred days so centre-already-scheduled days come first (soft S3)
        days = _preferred_days(c)
        pref    = sorted([d for d in days if d in PREFERRED_DAYS],
                         key=lambda d: -centre_day[c.centre][d])
        nonpref = [d for d in days if d not in PREFERRED_DAYS]
        days    = pref + nonpref

        assigned = False
        for day in days:
            for slot1, slot2 in slot_options:
                slots_used = [slot1] + ([slot2] if slot2 else [])

                if any(s in teacher_busy[teacher][day] for s in slots_used):
                    continue

                for room in _rooms_by_fit(rooms, c.students, preferred_prefix):
                    if any(s in room_busy[room.code][day] for s in slots_used):
                        continue

                    # ✓ Valid assignment found
                    for s in slots_used:
                        teacher_busy[teacher][day].add(s)
                        room_busy[room.code][day].add(s)
                    centre_day[c.centre][day] += 1

                    scheduled.append(ScheduledClass(
                        entry=c, day=day, slot1=slot1, slot2=slot2,
                        room=room.code, lecturer1=teacher,
                    ))
                    assigned = True
                    break

                if assigned:
                    break
            if assigned:
                break

        if not assigned:
            unscheduled.append(c.code)

    return scheduled, unscheduled


# ─── Phase 4 — Write Excel output ─────────────────────────────────────────────

def write_output_wb(wb, scheduled: list):
    """Fill an already-loaded workbook with scheduled classes (in-place)."""

    # ── 4a. Fill "Class list answer" sheet ──────────────────────────────────
    ws = wb["Class list answer"]

    # Clear existing data rows (leave row 1 header + formula cells untouched)
    for row in ws.iter_rows(min_row=2, max_row=200):
        for cell in row:
            if cell.value is not None:
                v = str(cell.value)
                if not v.startswith("="):
                    cell.value = None

    # Write scheduled classes, sorted by class code
    for i, sc in enumerate(sorted(scheduled, key=lambda s: s.entry.code)):
        r = i + 2
        ws.cell(r, 1,  i + 1)                  # row number
        ws.cell(r, 2,  sc.entry.code)           # Course Code
        ws.cell(r, 3,  sc.entry.subject_name)   # English name
        ws.cell(r, 4,  sc.entry.subject_cn)     # Chinese name
        ws.cell(r, 5,  sc.lecturer1)            # Lecturer1
        ws.cell(r, 6,  sc.lecturer2)            # Lecturer2 (Net)
        ws.cell(r, 7,  sc.lecturer3)            # Lecturer3 (Net)
        ws.cell(r, 8,  30)                      # Checking
        ws.cell(r, 9,  sc.day)                  # Day
        ws.cell(r, 10, sc.slot1)                # Time 1
        ws.cell(r, 11, sc.slot2)                # Time 2 (None for 2-hr)
        ws.cell(r, 12, sc.room)                 # Venue

    # ── 4b. Fill 5 daily timetable sheets ────────────────────────────────────
    # Fast lookup: (day, room, slot) -> ScheduledClass
    slot_lookup: dict = {}
    for sc in scheduled:
        slot_lookup[(sc.day, sc.room, sc.slot1)] = sc
        if sc.slot2:
            slot_lookup[(sc.day, sc.room, sc.slot2)] = sc

    for day, sheet_name in DAY_TO_SHEET.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found — skipping {day}")
            continue
        ws = wb[sheet_name]

        # Build room -> row map from column A (room data rows only, max row 37)
        room_row_map: dict = {}
        for row in ws.iter_rows(min_row=3, max_row=37, max_col=1):
            val = row[0].value
            if val and isinstance(val, str) and " - " in val:
                room_row_map[val] = row[0].row

        # Fill each slot group
        for slot, start_col in SLOT_TO_START_COL.items():
            for room_code, row_num in room_row_map.items():
                sc = slot_lookup.get((day, room_code, slot))
                if sc:
                    ws.cell(row_num, start_col,     sc.entry.code)
                    ws.cell(row_num, start_col + 1, sc.entry.students)
                    ws.cell(row_num, start_col + 2, sc.entry.subject_cn)
                    ws.cell(row_num, start_col + 3, sc.lecturer1)
                    # start_col + 4 = Time col: leave blank (standard slot)
                else:
                    # Clear any pre-existing non-formula data
                    for offset in range(4):
                        cell = ws.cell(row_num, start_col + offset)
                        if (cell.value is not None and
                                not str(cell.value).startswith("=")):
                            cell.value = None



def write_output(input_path: str, output_path: str, scheduled: list):
    """CLI helper: open template from disk, fill it, save to disk."""
    wb = openpyxl.load_workbook(input_path)
    write_output_wb(wb, scheduled)
    wb.save(output_path)
    print(f"\n  Saved: {output_path}")


# ─── Web API helpers ───────────────────────────────────────────────────────────

def collect_stats(scheduled: list, rooms: list, unscheduled: list) -> dict:
    """Return validation results as a plain dict (used by web API)."""
    room_cap = {r.code: r.capacity for r in rooms}
    violations: list = []
    teacher_slots: dict = defaultdict(list)
    room_slots:    dict = defaultdict(list)

    for sc in scheduled:
        slots = [sc.slot1] + ([sc.slot2] if sc.slot2 else [])
        for s in slots:
            teacher_slots[sc.lecturer1].append((sc.day, s, sc.entry.code))
            room_slots[sc.room].append((sc.day, s, sc.entry.code))
        cap = room_cap.get(sc.room, 0)
        if cap < sc.entry.students:
            violations.append(
                f"Overflow: {sc.entry.code} ({sc.entry.students} students)"
                f" in {sc.room} (cap {cap})"
            )

    for teacher, entries in teacher_slots.items():
        seen: dict = {}
        for day, slot, code in entries:
            k = (day, slot)
            if k in seen:
                violations.append(
                    f"Teacher clash: {teacher} - {seen[k]} & {code} on {day} {slot}"
                )
            else:
                seen[k] = code

    for room, entries in room_slots.items():
        seen = {}
        for day, slot, code in entries:
            k = (day, slot)
            if k in seen:
                violations.append(
                    f"Room clash: {room} - {seen[k]} & {code} on {day} {slot}"
                )
            else:
                seen[k] = code

    pref  = sum(1 for sc in scheduled if sc.day in PREFERRED_DAYS)
    total = len(scheduled)

    centre_dist: dict = defaultdict(lambda: defaultdict(int))
    for sc in scheduled:
        centre_dist[sc.entry.centre][sc.day] += 1

    return {
        "scheduled":     total,
        "total_classes": total + len(unscheduled),
        "unscheduled":   unscheduled,
        "violations":    violations,
        "preferred_pct": round(pref / total * 100) if total else 0,
        "centre_dist":   {c: dict(d) for c, d in centre_dist.items()},
    }


def run_from_bytes(excel_bytes: bytes) -> tuple:
    """
    Process an Excel file given as raw bytes.
    Returns (output_excel_bytes, stats_dict).
    Used by the web API — no filesystem access.
    """
    from io import BytesIO

    wb_read = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
    classes  = read_classes(wb_read)
    teachers = read_teachers(wb_read)
    rooms    = read_rooms(wb_read)

    assignments, no_teacher = assign_teachers(classes, teachers)
    scheduled, unscheduled  = schedule_classes(classes, assignments, rooms)

    # Load a fresh copy (without data_only) to preserve formulas in output
    wb_out = openpyxl.load_workbook(BytesIO(excel_bytes))
    write_output_wb(wb_out, scheduled)

    buf = BytesIO()
    wb_out.save(buf)
    output_bytes = buf.getvalue()

    stats = collect_stats(scheduled, rooms, unscheduled + list(no_teacher))
    return output_bytes, stats


# ─── Phase 5 — Validate and report ────────────────────────────────────────────

def validate_and_report(scheduled: list, rooms: list, unscheduled: list):
    room_cap = {r.code: r.capacity for r in rooms}

    teacher_slots: dict = defaultdict(list)
    room_slots:    dict = defaultdict(list)
    overflow:      list = []

    for sc in scheduled:
        slots = [sc.slot1] + ([sc.slot2] if sc.slot2 else [])
        for s in slots:
            teacher_slots[sc.lecturer1].append((sc.day, s, sc.entry.code))
            room_slots[sc.room].append((sc.day, s, sc.entry.code))
        cap = room_cap.get(sc.room, 0)
        if cap < sc.entry.students:
            overflow.append(
                f"  OVERFLOW  {sc.entry.code} "
                f"({sc.entry.students} students) → {sc.room} (cap {cap})"
            )

    teacher_clashes: list = []
    for teacher, entries in teacher_slots.items():
        seen: dict = {}
        for day, slot, code in entries:
            key = (day, slot)
            if key in seen:
                teacher_clashes.append(
                    f"  CLASH     Teacher {teacher} — "
                    f"{seen[key]} and {code} both on {day} {slot}"
                )
            else:
                seen[key] = code

    room_clashes: list = []
    for room, entries in room_slots.items():
        seen = {}
        for day, slot, code in entries:
            key = (day, slot)
            if key in seen:
                room_clashes.append(
                    f"  CLASH     Room {room} — "
                    f"{seen[key]} and {code} both on {day} {slot}"
                )
            else:
                seen[key] = code

    # Preferred-day usage
    pref_count = sum(1 for sc in scheduled if sc.day in PREFERRED_DAYS)
    total      = len(scheduled)
    pref_pct   = (pref_count / total * 100) if total else 0

    # Centre-grouping summary
    centre_day_dist: dict = defaultdict(lambda: defaultdict(int))
    for sc in scheduled:
        centre_day_dist[sc.entry.centre][sc.day] += 1

    sep = "=" * 62
    print(f"\n{sep}")
    print("  TIMETABLE SCHEDULER — VALIDATION REPORT")
    print(sep)
    print(f"\n  Classes scheduled   : {total}")
    print(f"  Classes unscheduled : {len(unscheduled)}")

    if unscheduled:
        print("\n  UNSCHEDULED (manual assignment needed):")
        for code in sorted(unscheduled):
            print(f"    - {code}")

    print(f"\n  Preferred-day usage (Mon/Tue/Thu): "
          f"{pref_count}/{total} ({pref_pct:.0f}%)")

    print("\n  Centre -> day distribution:")
    for centre in sorted(centre_day_dist):
        dist_str = ", ".join(
            f"{d[:3]}:{n}" for d, n in sorted(centre_day_dist[centre].items())
        )
        print(f"    {centre:<4}  {dist_str}")

    all_violations = teacher_clashes + room_clashes + overflow
    print(f"\n  Hard constraint violations:")
    if not all_violations:
        print("  OK  NONE - all hard constraints satisfied")
    else:
        for msg in all_violations:
            print(msg)

    status = "GO" if not all_violations and not unscheduled else "REVIEW NEEDED"
    print(f"\n  Overall status: {status}")
    print(sep)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: Input file not found:\n  {INPUT_FILE}")
        sys.exit(1)

    print(f"Input : {INPUT_FILE}")
    print(f"Output: {OUTPUT_FILE}\n")

    # Read with data_only so formula cells return their cached values
    wb_read = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    print("[1/5] Reading input data...")
    classes  = read_classes(wb_read)
    teachers = read_teachers(wb_read)
    rooms    = read_rooms(wb_read)
    print(f"      {len(classes)} classes  |  {len(teachers)} teachers  |  {len(rooms)} rooms")

    print("[2/5] Assigning teachers...")
    assignments, no_teacher = assign_teachers(classes, teachers)
    print(f"      {len(assignments)}/{len(classes)} classes assigned a lecturer")
    if no_teacher:
        print(f"      WARNING — no eligible teacher for: {no_teacher}")

    print("[3/5] Running greedy scheduler...")
    scheduled, unscheduled = schedule_classes(classes, assignments, rooms)
    print(f"      {len(scheduled)}/{len(classes)} classes scheduled")
    if unscheduled:
        print(f"      WARNING — could not schedule: {unscheduled}")

    print("[4/5] Writing Excel output...")
    write_output(INPUT_FILE, OUTPUT_FILE, scheduled)

    print("[5/5] Validating...")
    validate_and_report(scheduled, rooms, unscheduled)


if __name__ == "__main__":
    main()
