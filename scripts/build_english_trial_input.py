"""
build_english_trial_input.py — standalone English-only trial input.

Short-run isolation: a separate workbook containing ONLY the English (DAE102)
data, in the SAME format as the main planning file, run through the SAME engine.
No English-specific code path. Long-run merge = paste these sheets back into the
main planning file (already proven to schedule).

Derived from the cleaned main input so the availability / Net / name fixes carry
over. Output: data/input/Planning for Timetable — English.xlsx
"""
import os
import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(HERE, "data/input/Planning for Timetable (clean).xlsx")
DST  = os.path.join(HERE, "data/input/Planning for Timetable - English.xlsx")

ENG_TEACHERS = {
    "Ms. Lee Kit Wan", "Ms. Elise Ye", "Mr. Peter Barrett", "Ms. Cherry Ip",
    "Mr. Ray Leung", "Ms. Sasha Cheung", "Mr. Ivan Yuen", "Mr. Chris Hon",
}


def keep_rows(ws, keep_pred, header_rows=1):
    """Delete data rows (below header) where keep_pred(row_values) is False."""
    to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=header_rows + 1, values_only=True),
                            start=header_rows + 1):
        if not keep_pred(row):
            to_delete.append(i)
    for i in reversed(to_delete):
        ws.delete_rows(i, 1)


def main():
    wb = openpyxl.load_workbook(SRC)

    # Class list: keep only DAE102 (col A = Subject Code)
    keep_rows(wb["Class list"],
              lambda r: r and r[0] and str(r[0]).startswith("DAE102"))

    # Teacher load table: keep only English teachers (col B = name); header row 1
    keep_rows(wb["Teacher load table with subject"],
              lambda r: r and len(r) > 1 and r[1] and str(r[1]).strip() in ENG_TEACHERS)

    # Teacher Availability: keep only English teachers (col A = name); header row 1
    keep_rows(wb["Teacher Availability"],
              lambda r: r and r[0] and str(r[0]).strip() in ENG_TEACHERS)

    # Net Teachers, English Weekly, Centre Room Allocation, Class list answer: keep as-is.
    # (All rooms retained — harmless shared resource; English Weekly already blank.)

    # DAE102_CS7 is in Jo's arrangement but absent from the main planning file:
    # it's the "extra class in Oct" (police cadet, English-only, Net-exempt). Add
    # it here so the trial matches the arrangement's 27 classes 1:1.
    # ASSUMED: loading=4 and Student No=35 (copied from cadet classes CS1-3;
    # arrangement gives no student count for CS7).
    cl = wb["Class list"]
    cl.append(["DAE102_CS7", "English Language", "英國語文", None, None, None, 4, 35])
    print("+ added DAE102_CS7 (Oct extra cadet class; student count 35 = ASSUMED)")

    # Clear Class list answer so the engine runs TRUE V4 auto-assign (the clean
    # input already clears it, but the added CS7 must be auto-scheduled too).
    ws = wb["Class list answer"]
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    wb.save(DST)

    # Report
    wb2 = openpyxl.load_workbook(DST, data_only=True)
    n_cls = sum(1 for r in wb2["Class list"].iter_rows(min_row=2, values_only=True)
                if r and r[0])
    n_tch = sum(1 for r in wb2["Teacher load table with subject"].iter_rows(min_row=2, values_only=True)
                if r and len(r) > 1 and r[1])
    print(f"saved -> {DST}")
    print(f"  Class list: {n_cls} DAE102 classes")
    print(f"  Teacher load table: {n_tch} English teacher rows")
    print(f"  sheets: {wb2.sheetnames}")


if __name__ == "__main__":
    main()
