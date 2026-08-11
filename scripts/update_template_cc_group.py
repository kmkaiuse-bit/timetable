"""
Add a "CC Group" column + documentation to the V4 templates, in place.

For each template:
  1. Class list — append a "CC Group" header column (styled like the existing
     headers) and label the combine groups whose members are all present, as a
     live example of the feature (CS6+WT2, TW1+TW3, TS1+TS2 -> CC-DAE106-1/2/3).
  2. READ ME — append a "分班 / 規則 (2026-08)" section documenting the CC Group
     column, the DAE Mon/Tue/Thu day rule, and the loading soft cap.

Rollback: these files are versioned; discard the branch to revert.
"""
from copy import copy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "api" / "template_v4.xlsx",
    ROOT / "data" / "input" / "Planning for Timetable V4 Template.xlsx",
]

# Only groups whose members all exist in the templates (verified) — real examples.
EXAMPLE_LABELS = {
    "DAE106_CS6": "CC-DAE106-1", "DAE106_WT2": "CC-DAE106-1",
    "DAE106_TW1": "CC-DAE106-2", "DAE106_TW3": "CC-DAE106-2",
    "DAE106_TS1": "CC-DAE106-3", "DAE106_TS2": "CC-DAE106-3",
}

READ_ME_ROWS = [
    ("", ""),
    ("分班 / 規則 (2026-08)", ""),
    ("CC Group 欄",
     "要合併上課的班別填相同標籤（例：DAE106_CS6 與 DAE106_WT2 同填 CC-DAE106-1）；"
     "留空 = 不合併。只在 Class list answer 留空（自動排班）時生效。"),
    ("上課日",
     "DAE 科目只排 Mon / Tue / Thu（不排星期三、五）；Cadet 班維持 Mon / Wed / Fri。"),
    ("老師 Loading",
     "每週建議上限 6 班；可超過，只會提醒不會阻擋（超出者在畫面標示紅色「超限」）。"),
]


def update(path: Path) -> None:
    wb = openpyxl.load_workbook(path)

    # 1) Class list: add CC Group column
    cl = wb["Class list"]
    cc_col = cl.max_column + 1
    src = cl.cell(row=1, column=cl.max_column)          # style source = last header
    dst = cl.cell(row=1, column=cc_col, value="CC Group")
    dst.font = copy(src.font); dst.fill = copy(src.fill)
    dst.border = copy(src.border); dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    try:
        cl.column_dimensions[dst.column_letter].width = 14
    except Exception:
        pass

    labelled = 0
    for r in range(2, cl.max_row + 1):
        code = cl.cell(row=r, column=1).value
        if code and str(code).strip() in EXAMPLE_LABELS:
            cl.cell(row=r, column=cc_col, value=EXAMPLE_LABELS[str(code).strip()])
            labelled += 1

    # 2) READ ME: append section (skip if already present)
    rm = wb["READ ME"]
    existing = {str(rm.cell(row=r, column=1).value).strip()
                for r in range(1, rm.max_row + 1)}
    if "CC Group 欄" not in existing:
        label_style = copy(rm.cell(row=11, column=1).font)   # a notes-label cell
        start = rm.max_row + 1
        for i, (a, b) in enumerate(READ_ME_ROWS):
            ra = rm.cell(row=start + i, column=1, value=a or None)
            rm.cell(row=start + i, column=2, value=b or None)
            if a:
                ra.font = copy(label_style)

    wb.save(path)
    print(f"{path.name}: CC Group col at {dst.column_letter}, "
          f"labelled {labelled} example rows, READ ME updated")


if __name__ == "__main__":
    for f in FILES:
        if f.exists():
            update(f)
        else:
            print(f"MISSING: {f}")
