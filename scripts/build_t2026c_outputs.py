"""Build every T2026C deliverable Jo asked for.

    python scripts/build_t2026c_outputs.py <weekly_workbook.xlsx> <calendar.xlsx> [outdir]

Writes into data/output/ by default:

    Classes_DayTimeLocation_T2026C.xlsx   module A (paste-ready K:M + _Issues)
    DailyTimetable_T2026C.xlsx            module B
    ClassTimetable_T2026C.xlsx            module F
    DAE_TeachingSchedule_T2026C.xlsx      module E
"""

from __future__ import annotations

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "api"))

import english_net as en  # noqa: E402
import jo_export as jx  # noqa: E402

HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")

DAILY_COLUMNS = ["Raw", "Code", "Subject (in English)", "Subject (in Chinese)",
                 "Mode", "Topics", "Date", "Day", "Time", "Venue", "Lecturer",
                 "Programme", "Remark", "Year", "Month", "Date"]


def _header(ws, labels):
    for c, label in enumerate(labels, start=1):
        cell = ws.cell(1, c, label)
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


ACTION_FILL = PatternFill("solid", fgColor="FFE599")


def write_issues(wb, issues):
    """Anything Jo has to act on sorts to the top and is highlighted."""
    ws = wb.create_sheet("_Issues")
    _header(ws, ["Action needed", "Where to look (file / sheet / cell)",
                 "Kind", "Subject", "Detail"])
    ordered = sorted(issues, key=lambda i: (not i.get("action"), i["kind"],
                                            str(i["subject"])))
    for r, issue in enumerate(ordered, start=2):
        action = issue.get("action", "")
        ws.cell(r, 1, action)
        ws.cell(r, 2, issue.get("where", ""))
        ws.cell(r, 3, issue["kind"])
        ws.cell(r, 4, issue["subject"])
        ws.cell(r, 5, issue["detail"])
        for c in range(1, 6):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=(c == 5))
            if action:
                ws.cell(r, c).fill = ACTION_FILL
        if action:
            ws.cell(r, 1).font = Font(bold=True)
    for col, width in zip("ABCDE", (34, 46, 16, 22, 90)):
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A1:E{max(1, len(ordered) + 1)}"


def write_module_a(path, class_rows, issues):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classes K-M"
    _header(ws, ["Classes row", "Subject Code", "Day", "Time", "Location",
                 "Room code (grid)"])
    for i, rec in enumerate(class_rows, start=2):
        ws.cell(i, 1, rec["row"])
        ws.cell(i, 2, rec["code"])
        ws.cell(i, 3, rec.get("day"))
        ws.cell(i, 4, rec.get("time"))
        ws.cell(i, 5, rec.get("location"))
        ws.cell(i, 6, rec.get("room"))
    for col, width in zip("ABCDEF", (12, 18, 12, 16, 16, 16)):
        ws.column_dimensions[col].width = width
    write_issues(wb, issues)
    wb.save(path)


def write_module_b(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DailyTimetable_T2026C"
    _header(ws, DAILY_COLUMNS)
    for i, rec in enumerate(rows, start=2):
        ws.cell(i, 1, rec["Raw"])
        ws.cell(i, 2, rec["Code"])
        ws.cell(i, 3, rec["Subject (in English)"])
        ws.cell(i, 4, rec["Subject (in Chinese)"])
        ws.cell(i, 5, rec["Mode"])
        ws.cell(i, 6, rec["Topics"])
        ws.cell(i, 7, f"=TEXT(DATE(N{i},O{i},P{i}),\"YYYY-MM-DD\")")
        ws.cell(i, 8, rec["Day"])
        ws.cell(i, 9, rec["Time"])
        ws.cell(i, 10, rec["Venue"])
        ws.cell(i, 11, rec["Lecturer"])
        ws.cell(i, 12, rec["Programme"])
        ws.cell(i, 13, rec["Remark"])
        ws.cell(i, 14, rec["Year"])
        ws.cell(i, 15, rec["Month"])
        ws.cell(i, 16, rec["DateNum"])
    for col, width in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
                           "K", "L", "M", "N", "O", "P"],
                          (6, 14, 40, 24, 6, 7, 12, 11, 13, 12, 20, 16, 22, 7, 7, 7)):
        ws.column_dimensions[col].width = width
    wb.save(path)


def write_module_c(path, result, issues):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ClassTeacher26-27 M-R"
    _header(ws, ["Class", "Subject Code", "Day", "Time", "Location"]
            + en.BLOCK_LABELS + ["Net hrs T1", "Net hrs T2"])
    classes, blocks = result["classes"], result["blocks"]
    for i, code in enumerate(sorted(classes), start=2):
        c = classes[code]
        ws.cell(i, 1, code.split("_", 1)[1])
        ws.cell(i, 2, code)
        ws.cell(i, 3, c["day"])
        ws.cell(i, 4, c["time"])
        ws.cell(i, 5, c["location"])
        for b in range(len(en.BLOCK_LABELS)):
            ws.cell(i, 6 + b, blocks[b][code])
        for t, rng in ((0, range(0, 3)), (1, range(3, 6))):
            hrs = sum(en.BLOCK_HOURS for b in rng
                      if blocks[b][code] in en.NET_TEACHERS)
            ws.cell(i, 12 + t, hrs)
    for col, width in zip("ABCDEFGHIJKLM",
                          (8, 16, 11, 14, 12, 20, 20, 20, 20, 20, 20, 11, 11)):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("TeacherLoading")
    _header(ws2, ["Teacher", "Blocks", "Hours", "Target hrs", "Difference",
                  "Net teacher", "Classes taught"])
    taught = {}
    for code in classes:
        for b in range(len(en.BLOCK_LABELS)):
            taught.setdefault(blocks[b][code], set()).add(code.split("_", 1)[1])
    for i, (teacher, n) in enumerate(
            sorted(result["used"].items(), key=lambda x: -x[1]), start=2):
        target = en.HOUR_BUDGET.get(teacher)
        ws2.cell(i, 1, teacher)
        ws2.cell(i, 2, n)
        ws2.cell(i, 3, n * en.BLOCK_HOURS)
        ws2.cell(i, 4, target)
        ws2.cell(i, 5, None if target is None else n * en.BLOCK_HOURS - target)
        ws2.cell(i, 6, "Yes" if teacher in en.NET_TEACHERS else "")
        ws2.cell(i, 7, ", ".join(sorted(taught.get(teacher, ()))))
    row = 2 + len(result["used"])
    ws2.cell(row + 1, 1, "Total").font = Font(bold=True)
    ws2.cell(row + 1, 2, sum(result["used"].values()))
    ws2.cell(row + 1, 3, sum(result["used"].values()) * en.BLOCK_HOURS)
    for col, width in zip("ABCDEFG", (24, 8, 8, 11, 11, 12, 70)):
        ws2.column_dimensions[col].width = width

    write_issues(wb, issues)
    wb.save(path)


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        return 1
    weekly_path, calendar_path = sys.argv[1], sys.argv[2]
    outdir = sys.argv[3] if len(sys.argv) > 3 else os.path.join("data", "output")
    os.makedirs(outdir, exist_ok=True)

    issues = []
    wb = openpyxl.load_workbook(weekly_path, data_only=True)

    classes = jx.read_classes(wb)
    room_map = jx.read_classroom_map(wb, issues, source=os.path.basename(weekly_path))
    weekly_name = os.path.basename(weekly_path)
    placements = jx.read_grid(wb, issues, {c["code"] for c in classes},
                              source=weekly_name)
    slots = jx.build_class_slots(placements, issues)
    class_rows = jx.module_a(classes, slots, room_map, issues,
                             source=weekly_name)
    wb.close()

    issues.extend(jx.validate(class_rows))

    english = en.assign(class_rows, issues)
    english_blocks = None
    if english:
        english_blocks = {code: [english["blocks"][b][code] for b in range(3)]
                          for code in english["classes"]}

    calendar = jx.read_calendar(calendar_path, issues,
                                source=os.path.basename(calendar_path))
    daily = jx.module_b(class_rows, calendar, issues, english_blocks)

    class_view = jx.module_f(class_rows, issues)
    teacher_view = jx.module_e(class_rows, issues)

    a_path = os.path.join(outdir, "Classes_DayTimeLocation_T2026C.xlsx")
    b_path = os.path.join(outdir, "DailyTimetable_T2026C.xlsx")
    f_path = os.path.join(outdir, "ClassTimetable_T2026C.xlsx")
    e_path = os.path.join(outdir, "DAE_TeachingSchedule_T2026C.xlsx")
    c_path = os.path.join(outdir, "EnglishNetAssignment_T2026C.xlsx")

    write_module_a(a_path, class_rows, issues)
    write_module_b(b_path, daily)
    write_issues(class_view, [i for i in issues if i["kind"] == "class view"])
    class_view.save(f_path)
    write_issues(teacher_view, [i for i in issues if i["kind"] == "teacher view"])
    teacher_view.save(e_path)
    if english:
        write_module_c(c_path, english,
                       [i for i in issues
                        if i["kind"] in ("net infeasible", "net structure",
                                         "unfilled block", "budget mismatch")])

    placed = sum(1 for r in class_rows if r.get("day"))
    print(f"Classes rows           : {len(class_rows)}")
    print(f"  with Day/Time/Location: {placed}")
    print(f"  left blank            : {len(class_rows) - placed}")
    print(f"Grid placements read   : {len(placements)}")
    print(f"Daily timetable rows   : {len(daily)}")
    todo = sum(1 for i in issues if i.get("action"))
    print(f"Issues logged          : {len(issues)}  "
          f"(needing Jo's decision: {todo})")
    print()
    outputs = [a_path, b_path, f_path, e_path] + ([c_path] if english else [])
    for path in outputs:
        print("wrote", path)
    if not english:
        print("!! English Net assignment could not be solved -- see _Issues")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
