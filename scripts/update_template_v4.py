"""
update_template_v4.py — bring api/template_v4.xlsx in line with the engine.

Adds the two optional English-trial sheets so Jo can build a valid English
input from the template, and corrects the stale "built-in defaults" note (the
hard-coded fallback was removed 2026-07-08 — a blank English Weekly now means
full auto-assignment).
"""
import os
import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TPL  = os.path.join(HERE, "api/template_v4.xlsx")

EW_HEADER = ["Class Code", "T2025C wk1-5", "T2025C wk6-10", "T2025C wk11-15",
             "T2026A wk1-5", "T2026A wk6-10", "T2026A wk11-15"]
EW_NOTE = ("# 選填：填寫則鎖定該班該 block 的老師；留空則系統自動分配（顧及可用時間、"
           "Net 要求、車程）。")


def main():
    wb = openpyxl.load_workbook(TPL)

    if "Net Teachers" not in wb.sheetnames:
        ws = wb.create_sheet("Net Teachers")
        ws["A1"] = "Teacher Name"
        ws["A2"] = "# 每行一個 Net teacher 姓名（需與 Teacher load table 一致）。"
        print("+ added 'Net Teachers' sheet")

    if "English Weekly" not in wb.sheetnames:
        ws = wb.create_sheet("English Weekly")
        for i, h in enumerate(EW_HEADER, 1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value=EW_NOTE)
        print("+ added 'English Weekly' sheet")

    # Update READ ME: correct English Weekly semantics + list the optional sheets
    ws = wb["READ ME"]
    last = ws.max_row
    ws.cell(row=last + 2, column=1, value="選填 Sheet（英文科 / Net teacher 試用）")
    ws.cell(row=last + 3, column=1, value="Net Teachers")
    ws.cell(row=last + 3, column=2,
            value="列出 Net teacher 姓名。系統確保每個非豁免英文班每學期 ≥1 個 Net block")
    ws.cell(row=last + 4, column=1, value="English Weekly")
    ws.cell(row=last + 4, column=2,
            value="⭐ 留空則系統自動分配英文老師；只在想手動鎖定個別班時才填")

    wb.save(TPL)
    print(f"saved -> {TPL}")
    print("sheets:", openpyxl.load_workbook(TPL).sheetnames)


if __name__ == "__main__":
    main()
