"""
build_clean_input.py — produce a corrected test input for the V4 auto-scheduler.

Fixes three data-quality problems in data/input/Planning for Timetable.xlsx:
  A. English teachers have blank availability -> fill Mon/Tue/Thu (AM/PM) from
     English Teacher Arrangement.xlsx rows 31-40.
  B. Honorific mismatch: "Mr. Cherry Ip" / "Mr. Lee Kit Wan" in the load table
     vs "Ms." elsewhere -> unify to "Ms." across every sheet.
  C. English Weekly sheet is pre-filled with Jo's answer -> clear it so the
     engine runs true auto-assignment instead of the preassigned override.

Output: data/input/Planning for Timetable (clean).xlsx
"""
import os
import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(HERE, "data/input/Planning for Timetable.xlsx")
DST  = os.path.join(HERE, "data/input/Planning for Timetable (clean).xlsx")

# ── Fix B: name unifications (wrong -> correct), applied to every cell ─────────
NAME_FIX = {
    "Mr. Cherry Ip":   "Ms. Cherry Ip",
    "Mr. Lee Kit Wan": "Ms. Lee Kit Wan",
}

# ── Fix A: English teacher availability, from arrangement rows 31-40 ───────────
# Value = set of available days with AM/PM. Anything not listed = unavailable (N).
# AM -> 0900,1100 ; PM -> 1400,1600
AVAIL = {
    "Ms. Lee Kit Wan":   {"Thu": "AM"},
    "Ms. Elise Ye":      {"Tue": "PM", "Thu": "PM"},
    "Mr. Peter Barrett": {"Mon": "AMPM", "Tue": "AMPM", "Thu": "AMPM"},
    "Ms. Cherry Ip":     {"Mon": "AMPM", "Tue": "AMPM", "Thu": "AMPM"},
    "Mr. Ray Leung":     {"Tue": "AMPM"},
    "Ms. Sasha Cheung":  {"Mon": "AMPM", "Tue": "AMPM", "Thu": "AMPM"},
    "Mr. Ivan Yuen":     {"Mon": "AMPM", "Thu": "AMPM"},
    "Mr. Chris Hon":     {"Mon": "AMPM", "Tue": "AMPM", "Thu": "AMPM"},
}
DAYS  = ["Mon", "Tue", "Wed", "Thu", "Fri"]
SLOTS = ["0900", "1100", "1400", "1600"]


def is_available(avail_days, day, slot):
    spec = avail_days.get(day)
    if not spec:
        return False
    if spec == "AMPM":
        return True
    if spec == "AM":
        return slot in ("0900", "1100")
    if spec == "PM":
        return slot in ("1400", "1600")
    return False


def main():
    wb = openpyxl.load_workbook(SRC)

    # ── Fix B: rename across all sheets/cells ─────────────────────────────────
    renamed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value.strip()
                    if v in NAME_FIX:
                        cell.value = NAME_FIX[v]
                        renamed += 1
    print(f"[B] renamed {renamed} cells (Mr.->Ms. Cherry Ip / Lee Kit Wan)")

    # ── Fix A: fill Teacher Availability for the 8 English teachers ───────────
    ws = wb["Teacher Availability"]
    header = [c.value for c in ws[1]]           # "Teacher","Mon 0900",...
    col_of = {}
    for i, h in enumerate(header):
        if h and " " in str(h):
            d, s = str(h).split()
            col_of[(d, s)] = i
    filled_rows = 0
    for row in ws.iter_rows(min_row=2):
        name = row[0].value
        if not name:
            continue
        name = str(name).strip()
        if name not in AVAIL:
            continue
        for (d, s), ci in col_of.items():
            mark = None if is_available(AVAIL[name], d, s) else "N"
            row[ci].value = mark
        filled_rows += 1
    print(f"[A] filled availability for {filled_rows} English teachers")

    # ── Fix C: clear English Weekly data rows (keep header + note) ─────────────
    ws = wb["English Weekly"]
    cleared = 0
    for row in ws.iter_rows(min_row=2):
        first = row[0].value
        if first and str(first).strip().startswith("DAE"):
            for cell in row:
                cell.value = None
            cleared += 1
    print(f"[C] cleared {cleared} English Weekly pre-assigned rows")

    # ── Fix D: clear Class list answer so the engine runs TRUE V4 auto-assign ──
    # A populated answer makes run_v4_from_bytes replay Jo's pre-filled
    # Day/Time/Room (V3) instead of scheduling from scratch. Empty = full auto.
    ws = wb["Class list answer"]
    n = ws.max_row
    if n >= 2:
        ws.delete_rows(2, n - 1)
    print(f"[D] cleared Class list answer ({n - 1} rows) -> true V4 auto-assign")

    wb.save(DST)
    print(f"saved -> {DST}")


if __name__ == "__main__":
    main()
