"""
Populate Jo's DAE106 combine-course groupings into a COPY of the input workbook.

Source of the groupings (Jo, 2026-08):
    CC-DAE106-1 : DAE106_CS6 + DAE106_WT2                 (centre CSW)
    CC-DAE106-2 : DAE106_TW1 + DAE106_TW3                 (centre TW)
    CC-DAE106-3 : DAE106_TS1 + DAE106_TS2                 (centre TM)
    CC-DAE106-4 : DAE106_TS3 + DAE106_TS4 + DAE106_TM3    (centre TM)
    CC-DAE106-5 : DAE106_KT3 + DAE106_KT4                 (centre KT)

What this script does (writes to a NEW file — original is never modified):
  1. Copy the source workbook to a new path.
  2. Add a "CC Group" column to the "Class list" sheet (the scheduler detects any
     header containing both "cc" and "group" — see _load_classes()).
  3. Write the CC-DAE106-N label onto every combine class row that already exists.
  4. Append rows for the two combine classes missing from the sheet
     (DAE106_TM3, DAE106_KT4) using a sibling DAE106 row as a template for the
     subject names / loading, and Jo's student counts.

Student counts on Jo's sheet differ from the workbook for several classes; this
script does NOT touch existing student numbers — that discrepancy is reported
separately for Jo to confirm.
"""
import shutil
import sys
from pathlib import Path

import openpyxl

# Source is the CLEAN (empty Class-list answer) workbook: the CC-combine phase only
# runs in auto-assign mode (n_sched == 0). Putting the CC column into the pre-filled
# production file has no effect, because its pre-placed classes exhaust the L1/L2
# day/centre search. Use the auto-assign source so the combine actually schedules.
ROOT = Path(__file__).resolve().parents[1]
SRC  = ROOT / "data" / "input" / "Planning for Timetable (clean).xlsx"
DST  = ROOT / "data" / "input" / "Planning for Timetable (CC combine).xlsx"

# code -> CC group label
COMBINE = {
    "DAE106_CS6": "CC-DAE106-1", "DAE106_WT2": "CC-DAE106-1",
    "DAE106_TW1": "CC-DAE106-2", "DAE106_TW3": "CC-DAE106-2",
    "DAE106_TS1": "CC-DAE106-3", "DAE106_TS2": "CC-DAE106-3",
    "DAE106_TS3": "CC-DAE106-4", "DAE106_TS4": "CC-DAE106-4", "DAE106_TM3": "CC-DAE106-4",
    "DAE106_KT3": "CC-DAE106-5", "DAE106_KT4": "CC-DAE106-5",
}
# Classes that do not yet exist as rows — append them. Student counts from Jo's sheet.
MISSING = {"DAE106_TM3": 8, "DAE106_KT4": 9}

TEMPLATE_CODE = "DAE106_TS3"   # sibling row to clone subject-name / loading columns from


def main() -> int:
    if not SRC.exists():
        print(f"SOURCE NOT FOUND: {SRC}")
        return 1
    shutil.copyfile(SRC, DST)

    wb = openpyxl.load_workbook(DST)
    ws = wb["Class list"]

    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    cc_idx = next((i for i, h in enumerate(header)
                   if "cc" in h.lower() and "group" in h.lower()), None)
    if cc_idx is None:
        cc_idx = ws.max_column          # 0-based index of the new column
        ws.cell(row=1, column=cc_idx + 1, value="CC Group")

    # Grab template row values (subject names, loading) for the appended rows
    template = None
    row_by_code = {}
    for row in ws.iter_rows(min_row=2):
        code = row[0].value
        if not code:
            continue
        code = str(code).strip()
        row_by_code[code] = row
        if code == TEMPLATE_CODE:
            template = [c.value for c in row]

    labelled, appended = [], []

    # 1) Label existing rows
    for code, label in COMBINE.items():
        if code in row_by_code:
            ws.cell(row=row_by_code[code][0].row, column=cc_idx + 1, value=label)
            labelled.append(code)

    # 2) Append missing rows
    for code, students in MISSING.items():
        if code in row_by_code:
            continue
        new_row = [None] * (cc_idx + 1)
        if template:
            for i in range(min(len(template), cc_idx)):
                new_row[i] = template[i]
        new_row[0] = code                       # class code
        new_row[7] = students                   # Student No
        new_row[8] = new_row[9] = new_row[10] = None   # clear venue/time/date
        new_row[cc_idx] = COMBINE[code]         # CC Group label
        ws.append(new_row)
        appended.append(f"{code} (students={students})")

    wb.save(DST)

    print(f"WROTE: {DST}")
    print(f"CC Group column at 0-based index {cc_idx} (spreadsheet col {cc_idx + 1})")
    print(f"Labelled {len(labelled)} existing rows: {sorted(labelled)}")
    print(f"Appended {len(appended)} new rows: {appended}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
