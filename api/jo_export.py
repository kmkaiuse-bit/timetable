"""T2026C export pipeline for Jo's six deliverables.

Reads the weekly timetable workbook that Jo maintains by hand and produces:

  A  Day / Time / Location for every row of the `Classes` sheet
  B  the daily timetable (15 lesson days expanded into one row per session)
  E  a teacher-view grid per teaching staff member
  F  a class-view grid per class

Nothing here re-schedules anything -- Jo has already placed every class on the
`Mon(Term1)` .. `Sat(Term1)` sheets and these modules only read that placement
back out.  See docs/plans/2026-08-18-001-feat-jo-t2026c-deliverables-plan.md.
"""

from __future__ import annotations

import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# configuration
# --------------------------------------------------------------------------

TERM1_SHEET_RE = re.compile(r"^(Mon|Tue|Wed|Thu|Fri|Sat)\(Term1\)\s*$")

DAY_FULL = {"Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday",
            "Thu": "Thursday", "Fri": "Friday", "Sat": "Saturday"}
DAY_ABBR = {v: k for k, v in DAY_FULL.items()}
DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

# Agreed with Jo 2026-08-18: police college handles the first six separately and
# DAE270 waits on AASFP, so they carry no placement this round.
EXCLUDED_CODES = {"DAE256_G", "DAE256_H", "DAE258_G", "DAE258_H",
                  "DAE260_G", "DAE260_H", "DAE270"}

PROGRAMME = "DAE - FT2026C"
MODE = "D"
LESSON_DAYS = 15

# Jo confirmed each CSW room keeps its own code -- the fill-down in the
# "Classroom Code" sheet that maps C1..C6 all to "CSW - C1" is a mistake.
CLASSROOM_MAP_OVERRIDE = {f"CSW - C{i}": f"CSW - C{i}" for i in range(1, 7)}

# Short subject names as they appear in Jo's class-view sample.
SUBJECT_SHORT = {"DAE101": "中文", "DAE102": "英文"}

TEACHER_VIEW_STAFF = [
    ("Jo", "Ms. Jo Hugh"),
    ("Kevin", "Mr. Kevin Ho"),
    ("Chris", "Mr. Chris Chau"),
    ("Man", "Mr. Man Li"),
    ("Alan", "Mr. Alan Ho"),
    ("Eddie", "Mr. Eddie Cheung"),
    ("Edward", "Mr. Edward Siu"),
    ("Cherry", "Ms. Cherry Ip"),
    ("Loveleen", "Ms. Loveleen Kaur"),
    ("Hailey", "Ms. Hailey Wong"),
]

CLASS_VIEW_CLASSES = [
    "CS1", "CS2", "CS3", "CS4", "CS5", "CS6", "CS7",
    "WT1", "WT2", "SW1", "TW1", "TW2", "TW3",
    "TM1", "TM2", "TM3", "TM4", "TM5",
    "TS1", "TS2", "TS3", "TS4",
    "KT1", "KT2", "KT3", "KT4", "TK1",
    "ST1", "ST2", "FL1", "FL2", "FL3",
]

GRID_START = 8 * 60 + 30          # 0830
CLASS_VIEW_END = 18 * 60 + 30     # 1830
TEACHER_VIEW_END = 19 * 60 + 30   # 1930
SLOT_MINUTES = 30

_HONORIFIC_RE = re.compile(r"^(mr|ms|mrs|miss|dr|prof)\.?\s+", re.I)
# Jo's own spellings drift between files.
NAME_ALIASES = {"eddiechueng": "eddiecheung", "kaurloveleen": "loveleenkaur"}


def Issue(kind, subject, detail, action="", where=""):
    """One row of the `_Issues` report.

    `action` is non-empty when Jo has to decide something rather than just be
    told; those rows sort to the top and are highlighted.  `where` points at the
    exact cell to open -- "file / sheet / cell".
    """
    return {"kind": kind, "subject": subject, "detail": detail,
            "action": action, "where": where}


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------

def norm_name(name):
    """Fold a teacher name to a comparable key: no honorific, no punctuation."""
    if not name:
        return ""
    s = _HONORIFIC_RE.sub("", str(name).strip())
    s = re.sub(r"[^a-z]", "", s.lower())
    return NAME_ALIASES.get(s, s)


def parse_hhmm(text):
    """'0900' -> 540.  Returns None when the token is not a time."""
    if text is None:
        return None
    m = re.fullmatch(r"(\d{1,2})[:.]?(\d{2})", str(text).strip())
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if h > 23 or mi > 59:
        return None
    return h * 60 + mi


def fmt_hhmm(minutes):
    return f"{minutes // 60:02d}{minutes % 60:02d}"


def parse_range(text, issues=None, subject="", where=""):
    """Parse the many spellings of '0900 - 1100' into (start, end) minutes."""
    if text is None:
        return None
    raw = str(text).strip()
    m = re.fullmatch(r"(\d{3,4})\s*-\s*(\d{3,4})", raw)
    if not m:
        return None
    start, end = parse_hhmm(m.group(1)), parse_hhmm(m.group(2))
    if start is None or end is None or end <= start:
        return None
    # Jo has one '0900-1101' typo; snap odd minutes back to the half hour.
    if end % SLOT_MINUTES:
        fixed = int(round(end / SLOT_MINUTES) * SLOT_MINUTES)
        if issues is not None:
            issues.append(Issue("time typo", subject,
                                f"'{raw}' end time snapped to {fmt_hhmm(fixed)}",
                                where=where))
        end = fixed
    return start, end


def centre_of(room_code):
    """'KT - KT1' -> 'KT'."""
    return str(room_code).split("-")[0].strip() if room_code else ""


def class_suffix(code):
    """'DAE102_CS1' -> 'CS1'."""
    return code.split("_", 1)[1] if "_" in code else ""


def split_sessions(start, end):
    """A 4-hour placement is two sessions; a 2-hour placement is one."""
    n = max(1, int(round((end - start) / 120)))
    step = (end - start) / n
    return [(int(start + i * step), int(start + (i + 1) * step)) for i in range(n)]


# --------------------------------------------------------------------------
# readers
# --------------------------------------------------------------------------

def read_classes(wb):
    """The `Classes` sheet: one record per subject-class."""
    ws = wb["Classes"]
    out = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 2).value
        if not code or not str(code).strip():
            continue
        out.append({
            "row": r,
            "term": ws.cell(r, 1).value,
            "code": str(code).strip(),
            "subject_en": ws.cell(r, 3).value,
            "subject_ch": ws.cell(r, 4).value,
            "lecturers": [ws.cell(r, c).value for c in (5, 6, 7)
                          if ws.cell(r, c).value],
            "loading": ws.cell(r, 8).value,
            "students": ws.cell(r, 9).value,
        })
    return out


def read_classroom_map(wb, issues, source=""):
    """Room code -> the name used on the daily timetable."""
    ws = wb["Classroom Code"]
    prefix = f"{source} / " if source else ""
    mapping, row_of = {}, {}
    for r in range(2, ws.max_row + 1):
        code, daily = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not code:
            continue
        code = str(code).strip()
        mapping[code] = str(daily).strip() if daily else code
        row_of[code] = r
    for code, correct in CLASSROOM_MAP_OVERRIDE.items():
        if code in mapping and mapping[code] != correct:
            issues.append(Issue(
                "classroom map", code,
                f"the sheet maps this to '{mapping[code]}', which looks like a "
                f"fill-down error; using '{correct}' as you confirmed",
                where=f"{prefix}Classroom Code / B{row_of[code]}"))
            mapping[code] = correct
    return mapping


def read_grid(wb, issues, known_codes=None, source=""):
    """Every placement on the six Term 1 day sheets.

    `known_codes` limits the result to codes that exist on the `Classes` sheet;
    the day grids also carry Higher Diploma bookings that are not ours to export.
    """
    placements = []
    ignored = set()
    inferred_seen = set()
    prefix = f"{source} / " if source else ""
    for ws in wb.worksheets:
        m = TERM1_SHEET_RE.match(ws.title)
        if not m:
            continue
        day = m.group(1)

        header_row = None
        for r in range(1, 6):
            if any(isinstance(ws.cell(r, c).value, str)
                   and ws.cell(r, c).value.strip() == "Code"
                   for c in range(1, ws.max_column + 1)):
                header_row = r
                break
        if header_row is None:
            issues.append(Issue("grid", ws.title, "no 'Code' header row found",
                                where=f"{prefix}{ws.title}"))
            continue
        sheet_ref = f"{prefix}{ws.title.strip()}"

        block_cols = [c for c in range(1, ws.max_column + 1)
                      if isinstance(ws.cell(header_row, c).value, str)
                      and ws.cell(header_row, c).value.strip() == "Code"]

        # Row 1 carries the column-header time for blocks whose Time cell is blank.
        header_time = {}
        if header_row > 1:
            for c in block_cols:
                for probe in range(c, max(0, c - 5), -1):
                    t = parse_hhmm(ws.cell(header_row - 1, probe).value)
                    if t is not None:
                        header_time[c] = t
                        break

        current_room = None
        for r in range(header_row + 1, ws.max_row + 1):
            room_cell = ws.cell(r, 1).value
            room_here = str(room_cell).strip() if room_cell and str(room_cell).strip() else None
            if room_here:
                current_room = room_here
            for c in block_cols:
                code = ws.cell(r, c).value
                if not code or not str(code).strip():
                    continue
                code = str(code).strip()
                if known_codes is not None and code not in known_codes:
                    ignored.add(code)
                    continue
                code_cell = f"{get_column_letter(c)}{r}"
                time_cell = f"{get_column_letter(c + 4)}{r}"
                span = parse_range(ws.cell(r, c + 4).value, issues,
                                   subject=f"{day} {code}",
                                   where=f"{sheet_ref} / {time_cell}")
                if span is None:
                    start = header_time.get(c)
                    if start is None:
                        issues.append(Issue(
                            "time missing", f"{day} {code}",
                            "no Time cell and no column header time",
                            where=f"{sheet_ref} / {time_cell}"))
                        continue
                    span = (start, start + 120)
                inferred = room_here is None
                if inferred:
                    # Combine class: shares the room of the block above (Jo, 2026-08-18).
                    if current_room is None:
                        issues.append(Issue(
                            "room missing", f"{day} {code}",
                            "blank classroom with no block above it",
                            where=f"{sheet_ref} / A{r}"))
                        continue
                    key = (code, day, current_room)
                    if key not in inferred_seen:
                        inferred_seen.add(key)
                        issues.append(Issue(
                            "room inferred", f"{day} {code}",
                            f"classroom cell is blank; taken from the block above "
                            f"as '{current_room}' (combine class)",
                            where=f"{sheet_ref} / A{r}"))
                placements.append({
                    "day": day, "row": r, "code": code,
                    "room": current_room, "start": span[0], "end": span[1],
                    "room_inferred": inferred,
                    "sheet_ref": sheet_ref, "code_cell": code_cell,
                    "room_cell": f"A{r}",
                })
    if ignored:
        issues.append(Issue("ignored entry", "Term 1 day sheets",
                            "not on the Classes sheet, skipped: "
                            + ", ".join(sorted(ignored)),
                            where=f"{prefix}Mon(Term1) .. Sat(Term1)"))
    return placements


def read_calendar(path, issues, source=""):
    """(day, 'AM'|'PM') -> the 15 lesson dates, from columns AA..AM."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["DAE Calendar 2026-2027"]
    prefix = f"{source} / " if source else ""

    header_row = None
    for r in range(1, 12):
        if str(ws.cell(r, 27).value).strip() == "Lesson":
            header_row = r
            break
    if header_row is None:
        raise ValueError("cannot find the 'Lesson' header in the calendar sheet")

    columns = {}
    for c in range(28, 40):
        label = ws.cell(header_row, c).value
        m = re.match(r"^(Mon|Tue|Wed|Thu|Fri|Sat)\s*\((AM|PM)\)$", str(label).strip())
        if m:
            columns[(m.group(1), m.group(2))] = c

    dates = {}
    for key, c in columns.items():
        col = []
        for i in range(LESSON_DAYS):
            v = ws.cell(header_row + 1 + i, c).value
            if hasattr(v, "year"):
                col.append(v)
            else:
                col.append(None)
                issues.append(Issue(
                    "calendar", f"{key[0]} ({key[1]})",
                    f"lesson {i + 1} is '{v}', so those daily-timetable rows "
                    "were left out",
                    action="Please supply the date once it is fixed",
                    where=f"{prefix}DAE Calendar 2026-2027 / "
                          f"{get_column_letter(c)}{header_row + 1 + i}"))
        dates[key] = col
    wb.close()
    return dates


# --------------------------------------------------------------------------
# module A -- Day / Time / Location
# --------------------------------------------------------------------------

def build_class_slots(placements, issues):
    """Collapse the grid into one Day / Time / Location per subject-class."""
    by_code = defaultdict(list)
    for p in placements:
        by_code[p["code"]].append(p)

    slots = {}
    for code, items in by_code.items():
        days = sorted({i["day"] for i in items}, key=DAY_ORDER.index)
        if len(days) > 1:
            issues.append(Issue("multi-day", code,
                                f"placed on {', '.join(days)} -- using {days[0]}"))
        items = [i for i in items if i["day"] == days[0]]

        rooms = sorted({i["room"] for i in items if i["room"]})
        if len(rooms) > 1:
            issues.append(Issue("multi-room", code,
                                f"placed in {', '.join(rooms)} -- using {rooms[0]}"))
        room = rooms[0] if rooms else None

        spans = sorted({(i["start"], i["end"]) for i in items})
        merged = [list(spans[0])]
        for start, end in spans[1:]:
            if start <= merged[-1][1]:
                merged[-1][1] = max(merged[-1][1], end)
            else:
                merged.append([start, end])
        if len(merged) > 1:
            pretty = ", ".join(f"{fmt_hhmm(a)}-{fmt_hhmm(b)}" for a, b in merged)
            issues.append(Issue("split placement", code,
                                f"non-contiguous blocks {pretty} -- using the first"))
        start, end = merged[0]

        anchor = items[0]
        slots[code] = {
            "day": days[0], "start": start, "end": end, "room": room,
            "room_inferred": any(i.get("room_inferred") for i in items),
            "sheet_ref": anchor.get("sheet_ref", ""),
            "code_cell": anchor.get("code_cell", ""),
            "room_cell": anchor.get("room_cell", ""),
        }
    return slots


def module_a(classes, slots, room_map, issues, source=""):
    """Attach Day / Time / Location to each Classes row."""
    prefix = f"{source} / " if source else ""
    rows = []
    for rec in classes:
        code = rec["code"]
        slot = slots.get(code)
        if slot is None:
            if code not in EXCLUDED_CODES:
                issues.append(Issue(
                    "not placed", code,
                    "no placement on any Term 1 day sheet",
                    where=f"{prefix}Classes / B{rec['row']}"))
            rows.append({**rec, "day": None, "time": None, "location": None})
            continue
        location = room_map.get(slot["room"], slot["room"])
        rows.append({
            **rec,
            "day": DAY_FULL[slot["day"]],
            "time": f"{fmt_hhmm(slot['start'])} - {fmt_hhmm(slot['end'])}",
            "location": location,
            "room": slot["room"],
            "room_inferred": slot.get("room_inferred", False),
            "sheet_ref": slot.get("sheet_ref", ""),
            "code_cell": slot.get("code_cell", ""),
            "room_cell": slot.get("room_cell", ""),
            "start": slot["start"],
            "end": slot["end"],
        })
    return rows


# --------------------------------------------------------------------------
# module B -- daily timetable
# --------------------------------------------------------------------------

def module_b(class_rows, calendar, issues, english_blocks=None):
    """Expand into one row per session.

    `english_blocks` maps a DAE102 code to its six five-week block teachers; the
    first three cover T2026C, so the Lecturer column changes every five lessons.
    """
    english_blocks = english_blocks or {}
    out = []
    raw = 0
    for rec in class_rows:
        if rec.get("day") is None:
            continue
        day_abbr = DAY_ABBR[rec["day"]]
        sessions = split_sessions(rec["start"], rec["end"])

        periods = {"AM" if s < 13 * 60 else "PM" for s, _ in sessions}
        if len(periods) > 1:
            am, pm = calendar.get((day_abbr, "AM")), calendar.get((day_abbr, "PM"))
            if am != pm:
                issues.append(Issue(
                    "am/pm split", rec["code"],
                    f"{rec['time']} spans AM and PM on {rec['day']}, "
                    "whose AM and PM lesson dates differ -- dated as AM",
                    action="Please confirm which lesson dates apply",
                    where=rec.get("sheet_ref", "")))
            period = "AM"
        else:
            period = periods.pop()

        dates = calendar.get((day_abbr, period))
        if not dates:
            issues.append(Issue("calendar", rec["code"],
                                f"no calendar column for {day_abbr} ({period})"))
            continue

        default_lecturer = rec["lecturers"][0] if rec["lecturers"] else None
        term_blocks = english_blocks.get(rec["code"])
        for lesson_no, date in enumerate(dates, start=1):
            if date is None:
                continue
            lecturer = default_lecturer
            if term_blocks:
                lecturer = term_blocks[min((lesson_no - 1) // 5, len(term_blocks) - 1)]
            for slot_no, (start, end) in enumerate(sessions, start=1):
                raw += 1
                out.append({
                    "Raw": raw,
                    "Code": rec["code"],
                    "Subject (in English)": rec["subject_en"],
                    "Subject (in Chinese)": rec["subject_ch"],
                    "Mode": MODE,
                    "Topics": (lesson_no - 1) * len(sessions) + slot_no,
                    "Date": None,          # filled with a formula on write
                    "Day": rec["day"],
                    "Time": f"{fmt_hhmm(start)} - {fmt_hhmm(end)}",
                    "Venue": rec["location"],
                    "Lecturer": lecturer,
                    "Programme": PROGRAMME,
                    "Remark": None,
                    "Year": date.year,
                    "Month": date.month,
                    "DateNum": date.day,
                })
    return out


# --------------------------------------------------------------------------
# modules E / F -- the two grid views
# --------------------------------------------------------------------------

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")


def _write_grid_sheet(ws, label_key, label_value, title, entries, end_minute):
    """Shared renderer: 30-minute rows x Monday..Saturday."""
    ws.cell(2, 1, label_key).font = Font(bold=True)
    ws.cell(2, 2, label_value).font = Font(bold=True)
    ws.cell(2, 3, title).font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=8)

    headers = ["From", "To", "Monday", "Tuesday", "Wednesday",
               "Thursday", "Friday", "Saturday"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(3, c, h)
        cell.font = Font(bold=True)
        cell.fill = _HEAD_FILL
        cell.alignment = _WRAP
        cell.border = _BORDER

    n_rows = (end_minute - GRID_START) // SLOT_MINUTES
    for i in range(n_rows):
        r = 4 + i
        start = GRID_START + i * SLOT_MINUTES
        ws.cell(r, 1, fmt_hhmm(start)).alignment = _WRAP
        ws.cell(r, 2, fmt_hhmm(start + SLOT_MINUTES)).alignment = _WRAP
        for c in range(1, 9):
            ws.cell(r, c).border = _BORDER

    day_col = {d: 3 + i for i, d in enumerate(DAY_ORDER)}

    for e in entries:
        col = day_col.get(e["day"])
        if col is None:
            continue
        top = 4 + max(0, (e["start"] - GRID_START) // SLOT_MINUTES)
        bottom = 4 + min(n_rows, (e["end"] - GRID_START) // SLOT_MINUTES) - 1
        if bottom < top:
            continue
        cell = ws.cell(top, col, e["text"])
        cell.alignment = _WRAP
        if bottom > top:
            ws.merge_cells(start_row=top, start_column=col,
                           end_row=bottom, end_column=col)

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 7
    for c in range(3, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20


def module_f(class_rows, issues):
    """Class view: one sheet per class, core subjects only."""
    placed = [r for r in class_rows if r.get("day")]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for cls in CLASS_VIEW_CLASSES:
        ws = wb.create_sheet(cls)
        entries = []
        for rec in placed:
            if class_suffix(rec["code"]) != cls:
                continue
            prefix = rec["code"].split("_", 1)[0]
            name = SUBJECT_SHORT.get(prefix) or rec["subject_ch"] or rec["subject_en"]
            entries.append({
                "day": DAY_ABBR[rec["day"]],
                "start": rec["start"], "end": rec["end"],
                "text": f"{name}\n{centre_of(rec['room'])}",
            })
        if not entries:
            issues.append(Issue("class view", cls, "no subject placed for this class"))
        _write_grid_sheet(ws, "Class", cls, "Class Timetable T2026C",
                          entries, CLASS_VIEW_END)
    return wb


def module_e(class_rows, issues):
    """Teacher view: one sheet per staff member on Jo's list."""
    placed = [r for r in class_rows if r.get("day")]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for short, full in TEACHER_VIEW_STAFF:
        ws = wb.create_sheet(short)
        key = norm_name(full)
        entries = []
        for rec in placed:
            if not any(norm_name(l) == key for l in rec["lecturers"]):
                continue
            entries.append({
                "day": DAY_ABBR[rec["day"]],
                "start": rec["start"], "end": rec["end"],
                "text": f"{rec['code']}\n({centre_of(rec['room'])})",
            })
        if not entries:
            issues.append(Issue("teacher view", full, "no class assigned in Term 1"))
        _write_grid_sheet(ws, "Teacher", full, "Teaching Schedule T2026C",
                          entries, TEACHER_VIEW_END)
    return wb


# --------------------------------------------------------------------------
# validation -- cross-check the extraction against the timetable's own rules
# --------------------------------------------------------------------------

def _overlaps(a, b):
    return a["start"] < b["end"] and b["start"] < a["end"]


def _pairs(records):
    for i in range(len(records)):
        for j in range(i + 1, len(records)):
            if _overlaps(records[i], records[j]):
                yield records[i], records[j]


def validate(class_rows):
    """Clashes in Jo's grid that survived the extraction.

    A combine class legitimately puts two codes in one room with one teacher, so
    those are filtered out; what is left needs a human to look at it.
    """
    placed = [r for r in class_rows if r.get("day")]
    findings = []

    def lecturer(rec):
        return norm_name(rec["lecturers"][0]) if rec["lecturers"] else ""

    by_room = defaultdict(list)
    by_teacher = defaultdict(list)
    by_class = defaultdict(list)
    for rec in placed:
        by_room[(rec["day"], rec["room"])].append(rec)
        if rec["lecturers"]:
            by_teacher[(rec["day"], lecturer(rec))].append(rec)
        suffix = class_suffix(rec["code"])
        if suffix in CLASS_VIEW_CLASSES:
            by_class[(rec["day"], suffix)].append(rec)

    def who(rec):
        name = rec["lecturers"][0] if rec["lecturers"] else "?"
        return f"{rec['code']} ({rec['time']}, {name})"

    def cell_of(rec):
        return f"{rec['sheet_ref']} / {rec['code_cell']}" if rec.get("sheet_ref") else ""

    for (day, room), recs in sorted(by_room.items(), key=lambda x: str(x[0])):
        for a, b in _pairs(recs):
            if lecturer(a) and lecturer(a) == lecturer(b):
                continue          # combine class -- one teacher, one room
            guessed = [r for r in (a, b) if r.get("room_inferred")]
            if guessed:
                g = guessed[0]
                findings.append(Issue(
                    "CONFIRM room", f"{day} {room}",
                    f"{who(a)} vs {who(b)}. The classroom cell for {g['code']} "
                    f"({g['sheet_ref']} / {g['room_cell']}) is blank, so the room was "
                    f"taken from the block above. The class it lands on has a "
                    f"different subject and a different teacher, so this is probably "
                    f"not a combine class -- most likely the classroom is missing.",
                    action=f"Fill in the classroom at {g['room_cell']}",
                    where=f"{g['sheet_ref']} / {g['room_cell']}"))
            else:
                findings.append(Issue(
                    "room clash", f"{day} {room}",
                    f"{who(a)} vs {who(b)} -- different teachers, "
                    "so not a combine class",
                    action="Please check",
                    where=cell_of(a) or cell_of(b)))

    for (day, _), recs in sorted(by_teacher.items(), key=lambda x: str(x[0])):
        for a, b in _pairs(recs):
            if a["room"] == b["room"]:
                continue          # combine class -- one teacher, one room
            name = a["lecturers"][0]
            findings.append(Issue(
                "teacher clash", f"{day} {name}",
                f"{a['code']} ({a['time']}) in {a['room']} "
                f"vs {b['code']} ({b['time']}) in {b['room']} "
                "-- one teacher cannot be in two rooms at once",
                action="Please check",
                where=" and ".join(x for x in (cell_of(a), cell_of(b)) if x)))

    for (day, cls), recs in sorted(by_class.items(), key=lambda x: str(x[0])):
        for a, b in _pairs(recs):
            findings.append(Issue(
                "class clash", f"{day} {cls}",
                f"{a['code']} ({a['time']}, {a['room']}) "
                f"vs {b['code']} ({b['time']}, {b['room']}) "
                "-- the same students would be in two subjects at once",
                action="Please check",
                where=" and ".join(x for x in (cell_of(a), cell_of(b)) if x)))

    return findings
