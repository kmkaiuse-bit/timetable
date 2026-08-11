"""
timetable_scheduler.py  (v6 — v2 meeting constraints)
======================================================
v3: Day/Time/Room fixed from Class list answer → assigns Lec1/2/3 only.
v4: Class list answer empty → auto-assigns Day/Time/Room/Lec1/2/3.
v6 changes (from scheduling-meeting-v2):
  H4  → soft cap (warn only, never block)
  H6  → TKO allows any slot with start ≥ 10:00 (was two fixed windows)
  H9  → CSW↔SSP and TW→CSW/SSP allowed cross-centre exceptions
  S1  → teacher can also declare preferred-not centres (new sheet)
  S2  → tiandi gap check applies to students only (never teachers)
  CC  → two-phase: Core first, then CC Combine (L1/L2/H5/S4)

Both modes share the same SQLite engine and teacher-assignment logic.
run_from_bytes()    → v3 (requires pre-filled Class list answer)
run_v4_from_bytes() → v4 (auto-assigns when Class list answer is empty)
"""

import os
import re
import sys
import sqlite3
from collections import defaultdict
from typing import Optional

import openpyxl

# ─── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE  = os.path.join(SCRIPT_DIR, "Planning for Timetable.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "Timetable_Output.xlsx")

# ─── Embedded SQLite schema ───────────────────────────────────────────────────

_SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE subjects (
    code TEXT PRIMARY KEY, name_en TEXT, name_cn TEXT, loading_hrs INTEGER);

CREATE TABLE rooms (
    code TEXT PRIMARY KEY, centre TEXT, capacity INTEGER);

CREATE TABLE class_groups (
    code TEXT PRIMARY KEY, centre TEXT, is_police_cadet INTEGER DEFAULT 0);

CREATE TABLE classes (
    code TEXT PRIMARY KEY, subject_code TEXT, group_code TEXT,
    student_count INTEGER DEFAULT 0,
    language TEXT DEFAULT NULL,
    cc_group TEXT DEFAULT NULL);

CREATE TABLE teacher_centre_preference (
    teacher_id INTEGER, centre TEXT, pref_type TEXT DEFAULT 'avoid',
    PRIMARY KEY (teacher_id, centre));

CREATE TABLE teachers (
    id INTEGER PRIMARY KEY, name TEXT UNIQUE,
    is_net INTEGER DEFAULT 0);

CREATE TABLE teacher_subjects (
    teacher_id INTEGER, subject_code TEXT,
    lec1_quota INTEGER DEFAULT 0, backup_quota INTEGER DEFAULT 0,
    PRIMARY KEY (teacher_id, subject_code));

CREATE TABLE teacher_unavailability (
    teacher_id INTEGER, day TEXT, start_time TEXT,
    PRIMARY KEY (teacher_id, day, start_time));

CREATE TABLE group_rooms (
    group_code TEXT PRIMARY KEY, room_code TEXT);

CREATE TABLE schedule (
    class_code TEXT PRIMARY KEY,
    group_code TEXT,
    day TEXT, time1 TEXT, time2 TEXT, room_code TEXT,
    teacher1_id INTEGER, teacher2_id INTEGER, teacher3_id INTEGER);
"""

# ─── Constants ────────────────────────────────────────────────────────────────

DAY_TO_SHEET = {
    "Monday":    "Mon(Term1) ",
    "Tuesday":   "Tue(Term1)",
    "Wednesday": "Wed(Term1)",
    "Thursday":  "Thu(Term1)",
    "Friday":    "Fri(Term1)",
}

_VENUE_ALIAS = {
    "WT - WT":  "WT - WT1",
    "KT - A":   "KT - KT1",  "KT - B": "KT - KT2",  "KT - C": "KT - KT3",
    "TM - A":   "TM - TM1",  "TM - B": "TM - TM2",  "TM - C": "TM - TM3",
    "TW - A":   "TW - TW1",  "TW - B": "TW - TW2",  "TW - C": "TW - TW3",
    "ST - A":   "ST - ST1",  "ST - B": "ST - ST2",  "ST - C": "ST - ST3",
    "TK - TK":  "TKO - TKO",
}

_SUBJECT_NAME_MAP = {
    "MathsPlus":               "Maths Plus",
    "Career and Life Learning": "Career and Life Planning",
}

_MARKER_TO_SLOT = {
    "0900": "0900 - 1100",  900:  "0900 - 1100",
    "1100": "1100 - 1300",  1100: "1100 - 1300",
    "1400": "1400 - 1600",  1400: "1400 - 1600",
    "1600": "1600 - 1800",  1600: "1600 - 1800",
}

# ─── v4 auto-scheduler constants ──────────────────────────────────────────────

# Group prefix → room centre (only exceptions; everything else maps 1:1)
# TK → TKO fixes the latent bug where _pick_room looked for centre 'TK'
# but all TKO rooms are stored as centre 'TKO'.
_GROUP_CENTRE_ALIAS = {"CS": "CSW", "TK": "TKO"}

# Day preference order for DAE subjects: Mon > Tue > Thu ONLY.
# DAE classes are NOT scheduled on Wednesday or Friday (2026-08 requirement).
# NOTE: Cadet classes are the explicit exception — they keep Mon/Wed/Fri and use
# _CADET_DAY_PRIORITY (see Phase 0 / phase0_schedule_cadets). Do not add Wed/Fri
# back here, and do not use _DAY_PRIORITY.index() on cadet days (would ValueError).
_DAY_PRIORITY = ["Monday", "Tuesday", "Thursday"]

# 4-hour blocks (two consecutive 2h slots on the same day).
# Only non-overlapping blocks — the double-booking LIKE check would miss
# overlaps between e.g. 0900-1100 and 1000-1200, causing ghost assignments.
_TIME_BLOCKS = [
    ("0900 - 1100", "1100 - 1300"),  # morning
    ("1400 - 1600", "1600 - 1800"),  # afternoon
]

_SINGLE_SLOTS = ["0900 - 1100", "1100 - 1300", "1400 - 1600", "1600 - 1800"]

# H6: TKO air-con cost — any slot with start ≥ 10:00 is allowed (v6: was two fixed windows)
_TKO_TIME_BLOCKS = [
    ("1000 - 1200", "1200 - 1400"),  # 10:00–14:00
    ("1400 - 1600", "1600 - 1800"),  # 14:00–18:00 (added v6)
    ("1500 - 1700", "1700 - 1900"),  # 15:00–19:00
]
_TKO_SINGLE_SLOTS = [
    "1000 - 1200", "1100 - 1300", "1200 - 1400",
    "1400 - 1600", "1500 - 1700", "1600 - 1800", "1700 - 1900",
]

# H4: soft weekly loading cap (warn if exceeded, never block)
_TEACHER_WEEKLY_SESSION_CAP = 6

# English Net teacher assignment constants (DAE102 only)
_ENG_SUBJECT_CODE        = "DAE102"
_ENG_NET_MIN_BLOCKS      = 1          # ≥1 Net block per term = 20 Net hrs (J9 assumed; file: Hours = blocks × 20)
_ENG_MAX_TRAVEL_MIN      = 30         # English teacher same-day travel cap (minutes)
_ENG_TERMS               = ["T2025C", "T2026A"]
_ENG_WEEK_BLOCKS         = ["wk1-5", "wk6-10", "wk11-15"]
_ENG_NET_EXEMPT_GROUPS   = {"CS1", "CS2", "CS3", "CS7"}   # no Net hours requirement

# English (DAE102) weekly assignments come from auto-assignment
# (assign_english_weekly Mode 2), or from the input workbook's "English Weekly"
# sheet when Jo fills it as an explicit override. The hard-coded copy of Jo's
# arrangement was removed 2026-07-08 — it silently overrode auto-assignment.
# Empty dict = no built-in fallback.
_ENG_WEEKLY_PREASSIGNED: dict = {}

# H8: max travel time (minutes) allowed for teacher to work two centres same day
# Cross-centre pairs within this threshold → soft warning; beyond → hard block
_H8_MAX_TRAVEL_MIN = 90

# H3 TS→TM exception: J1 assumption enabled (confirm with Jo — reverse by setting False)
_TS_TM_ENABLED = True

# Issue diagnostics: suggested actions per reason code (used by Issues panel + Excel sheet)
_SUGGESTED_ACTIONS = {
    "NO_ROOM_CAPACITY": "Increase room capacity at this centre or split the class into smaller groups",
    "SLOT_ROOM":        "Room exists but fully booked — check for scheduling conflicts",
    "SLOT_TEACHER":     "Room available but no teacher free — check teacher loading panel",
    "SLOT_H9":          "Student group has conflicting class on same day — cannot share that day",
    "SLOT_H6":          "No valid TKO time slot available (class must start at or after 10:00)",
    "TEACHER_NONE":     "Add a teacher with this subject qualification to the Teachers sheet",
    "TEACHER_CAPACITY": "All qualified teachers at weekly cap — add more or reduce other sessions",
    "TEACHER_AVAIL":    "All qualified teachers unavailable on the scheduled day",
    "TEACHER_H8":       "No teacher can reach this centre within 90 min from prior commitment",
    "H3_TS_TM":         "J1 assumption applied: TS→TM cross-centre assigned — confirm with Jo",
    "H3_FALLBACK":      "J8 assumption applied: class moved to fallback centre — confirm with Jo",
    "NET_SHORTFALL":    "Class shares a crowded slot and its centre is >30 min from the free Net teachers — reschedule to a lighter day/time, or accept no Net coverage. Adding Net teachers won't help if none can reach this centre at this slot.",
    "CADET_ERROR":      "Cadet class could not be placed on Mon/Wed/Fri",
}

# CC Combine: max travel time (minutes) from any group's home centre to the CC centre
# None = no distance filter (default until Jo confirms threshold, pending J2)
_CC_MAX_TRAVEL_MIN = None

# H9: cross-centre pairs that are allowed on the same day (student exceptions v6)
_H9_ALLOWED_CROSS_CENTRE = {
    frozenset({"CSW", "SSP"}),  # CSW ↔ SSP bidirectional
    frozenset({"TW",  "CSW"}),  # TW  ↔ CSW
    frozenset({"TW",  "SSP"}),  # TW  ↔ SSP
}

# Centre distance rank from SSP/CSW core (higher = closer)
# Used for S4 CC Combine tie-breaking and travel risk assessment
# WT=黃大仙, TS=青衣, TW=荃灣, ST=沙田, SW=上環, KT=觀塘, TKO=將軍澳, TM=屯門, FL=粉嶺
_CENTRE_RANK = {
    "SSP": 9, "CSW": 9,
    "WT":  8,
    "TS":  7, "TW": 7,
    "ST":  6, "SW":  6,
    "KT":  5,
    "TKO": 4,
    "TM":  2, "FL":  2,
}

# MTR/transport travel time in minutes between centres (symmetric, ±5 min)
# Based on: SSP/CSW as core; WT=20, TS/TW=25, ST/SW=30, KT=35, TKO=45, TM/FL=60
# TM↔FL via NT inland route (~60 min); SW cross-harbour distances corrected
_TRAVEL_TIME: dict = {
    "SSP": {"SSP":  0, "CSW":  8, "WT": 20, "KT": 35, "TKO": 45, "ST": 30, "FL": 60, "SW": 30, "TW": 25, "TS": 25, "TM": 60},
    "CSW": {"SSP":  8, "CSW":  0, "WT": 20, "KT": 35, "TKO": 43, "ST": 30, "FL": 60, "SW": 30, "TW": 25, "TS": 25, "TM": 58},
    "WT":  {"SSP": 20, "CSW": 20, "WT":  0, "KT": 15, "TKO": 25, "ST": 25, "FL": 48, "SW": 32, "TW": 38, "TS": 38, "TM": 62},
    "KT":  {"SSP": 35, "CSW": 35, "WT": 15, "KT":  0, "TKO": 15, "ST": 30, "FL": 52, "SW": 38, "TW": 42, "TS": 42, "TM": 68},
    "TKO": {"SSP": 45, "CSW": 43, "WT": 25, "KT": 15, "TKO":  0, "ST": 40, "FL": 55, "SW": 42, "TW": 52, "TS": 50, "TM": 72},
    "ST":  {"SSP": 30, "CSW": 30, "WT": 25, "KT": 30, "TKO": 40, "ST":  0, "FL": 15, "SW": 42, "TW": 40, "TS": 40, "TM": 58},
    "FL":  {"SSP": 60, "CSW": 60, "WT": 48, "KT": 52, "TKO": 55, "ST": 15, "FL":  0, "SW": 65, "TW": 55, "TS": 58, "TM": 60},
    "SW":  {"SSP": 30, "CSW": 30, "WT": 32, "KT": 38, "TKO": 42, "ST": 42, "FL": 65, "SW":  0, "TW": 25, "TS": 28, "TM": 65},
    "TW":  {"SSP": 25, "CSW": 25, "WT": 38, "KT": 42, "TKO": 52, "ST": 40, "FL": 55, "SW": 25, "TW":  0, "TS": 10, "TM": 32},
    "TS":  {"SSP": 25, "CSW": 25, "WT": 38, "KT": 42, "TKO": 50, "ST": 40, "FL": 58, "SW": 28, "TW": 10, "TS":  0, "TM": 30},
    "TM":  {"SSP": 60, "CSW": 58, "WT": 62, "KT": 68, "TKO": 72, "ST": 58, "FL": 60, "SW": 65, "TW": 32, "TS": 30, "TM":  0},
}

# Nearest centres per centre, sorted by travel time (used for J8 room overflow fallback)
# Only includes centres within practical travel range (≤65 min)
_NEAREST_CENTRES: dict = {
    "SSP": ["CSW", "WT", "TW", "TS", "ST", "SW"],        # core pair; WT 20min
    "CSW": ["SSP", "WT", "TW", "TS", "ST", "SW"],        # core pair; WT 20min
    "WT":  ["KT", "SSP", "CSW", "TKO", "ST"],            # KT/SSP 15-20min
    "KT":  ["TKO", "WT", "ST", "SSP", "CSW"],            # TKO/WT 15min
    "TKO": ["KT", "WT", "ST", "SSP", "CSW"],             # KT 15min; isolated east
    "ST":  ["FL", "WT", "KT", "TKO", "SSP", "CSW"],      # FL 15min (East Rail)
    "FL":  ["ST", "WT", "KT", "TKO", "TM"],              # ST 15min; TM via NT 60min
    "SW":  ["TW", "TS", "SSP", "CSW", "WT"],             # TW/TS 25-28min; cross-harbour
    "TW":  ["TS", "SSP", "CSW", "SW", "TM"],             # TS 10min; TM via West Rail 32min
    "TS":  ["TW", "TM", "SSP", "CSW", "SW"],             # TW 10min; TM 30min (West Rail)
    "TM":  ["TS", "TW", "FL", "SSP", "CSW"],             # TS 30min; FL via NT 60min
}


def _travel_mins(centre_a: str, centre_b: str) -> int:
    """Return estimated MTR travel time in minutes between two centres."""
    if centre_a == centre_b:
        return 0
    return _TRAVEL_TIME.get(centre_a, {}).get(centre_b, 999)


# Slot order within the day (for S2 天地堂 check)
_SLOT_HALF_DAY = {
    "0900 - 1100": "AM", "1000 - 1200": "AM",
    "1100 - 1300": "AM", "1200 - 1400": "AM",
    "1400 - 1600": "PM", "1500 - 1700": "PM",
    "1600 - 1800": "PM", "1700 - 1900": "PM",
}
# Numeric position within the day for gap calculation
_SLOT_START_HOUR = {
    "0900 - 1100": 9,  "1000 - 1200": 10,
    "1100 - 1300": 11, "1200 - 1400": 12,
    "1400 - 1600": 14, "1500 - 1700": 15,
    "1600 - 1800": 16, "1700 - 1900": 17,
}

# ─── Database helpers ─────────────────────────────────────────────────────────

def _new_db() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.executescript(_SCHEMA)
    return conn


def _resolve_venue(v: str) -> str:
    return _VENUE_ALIAS.get(str(v).strip(), str(v).strip())


def _extract_group(class_code: str) -> str:
    """'DAE101_CS1' -> 'CS1'"""
    parts = str(class_code).split("_")
    return parts[1] if len(parts) >= 2 else ""


def _extract_centre(class_code: str) -> str:
    """'DAE101_CS1' -> 'CSW', 'DAE101_TK1' -> 'TKO' (applies GROUP_CENTRE_ALIAS)"""
    g = _extract_group(class_code)
    prefix = re.sub(r"\d+$", "", g)
    return _GROUP_CENTRE_ALIAS.get(prefix, prefix)


# ─── Phase 1: Populate database from Excel ────────────────────────────────────

def _load_subjects(conn: sqlite3.Connection, wb):
    ws = wb["Class list"]
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, name_en, name_cn, *_, loading = row[0], row[1], row[2], row[6]
        if not code or str(code) in seen:
            continue
        subj_code = str(code).split("_")[0]
        if subj_code not in seen:
            hrs = int(loading) if isinstance(loading, (int, float)) else 4
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO subjects VALUES (?,?,?,?)",
                    (subj_code, str(name_en or "").strip(),
                     str(name_cn or "").strip(), hrs))
            except Exception:
                pass
            seen.add(subj_code)


def _load_rooms(conn: sqlite3.Connection, wb):
    ws = wb["Centre Room Allocation"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        # Col D = "Max no of seats" (approved capacity); col C = "No of Seats" (physical seats).
        # Use Max when available — e.g. FL Max=34 > No=29, SSP-303 Max=100 < No=111.
        cap_no  = row[2] if len(row) > 2 else None
        cap_max = row[3] if len(row) > 3 else None
        cap = cap_max if isinstance(cap_max, (int, float)) and cap_max > 0 else cap_no
        if not isinstance(cap, (int, float)) or cap <= 0:
            continue
        code_str = str(code).strip()
        centre   = code_str.split(" - ")[0].strip()
        conn.execute("INSERT OR IGNORE INTO rooms VALUES (?,?,?)",
                     (code_str, centre, int(cap)))


def _normalise_language(raw) -> Optional[str]:
    """Normalise language cell to 'C' (Chinese) or 'E' (English), or None."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    if s in ("C", "中", "中文", "CHINESE", "广东话", "廣東話", "CANTONESE"):
        return "C"
    if s in ("E", "英", "英文", "ENGLISH"):
        return "E"
    return None


def _load_classes(conn: sqlite3.Connection, wb):
    ws = wb["Class list"]
    # Detect header to find optional Language and CC Group columns
    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    lang_col = next((i for i, h in enumerate(header) if "lang" in h.lower()), None)
    cc_col   = next((i for i, h in enumerate(header)
                     if "cc" in h.lower() and "group" in h.lower()), None)

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        code_str  = str(code).strip()
        subj_code = code_str.split("_")[0]
        group     = _extract_group(code_str)
        centre    = _extract_centre(code_str)
        students  = int(row[7]) if isinstance(row[7], (int, float)) else 0
        is_pc     = 1 if group in ("CS1", "CS2", "CS3") else 0
        language  = _normalise_language(row[lang_col]) if lang_col is not None and lang_col < len(row) else None
        cc_group  = str(row[cc_col]).strip() if (cc_col is not None and cc_col < len(row) and row[cc_col]) else None

        conn.execute("INSERT OR IGNORE INTO class_groups VALUES (?,?,?)",
                     (group, centre, is_pc))
        conn.execute("INSERT OR IGNORE INTO classes VALUES (?,?,?,?,?,?)",
                     (code_str, subj_code, group, students, language, cc_group))


# ─── Teacher name normalization ───────────────────────────────────────────────
# The same person can appear under different honorifics across sheets — e.g.
# "Mr. Cherry Ip" in the load table vs "Ms. Cherry Ip" in the Net / English
# sheets. Matching on the honorific-stripped core keeps cross-sheet joins from
# silently splitting one teacher into two records.

_TITLE_RE = re.compile(r"^(mr|mrs|ms|miss|dr|prof)\.?\s+", re.IGNORECASE)


def _canon_core(name) -> str:
    """Lower-cased teacher name with any leading honorific removed."""
    if not name:
        return ""
    core = _TITLE_RE.sub("", str(name).strip())
    return " ".join(core.split()).lower()


def _teacher_id_by_name(conn: sqlite3.Connection, name) -> Optional[int]:
    """Resolve a teacher id, tolerating honorific differences. None if no match."""
    if not name:
        return None
    name = str(name).strip()
    row = conn.execute("SELECT id FROM teachers WHERE name = ?", (name,)).fetchone()
    if row:
        return row[0]
    core = _canon_core(name)
    if not core:
        return None
    for r in conn.execute("SELECT id, name FROM teachers").fetchall():
        if _canon_core(r[1]) == core:
            return r[0]
    return None


def _get_or_create_teacher(conn: sqlite3.Connection, name: str) -> int:
    """Return the id of an existing teacher (matched by core) or insert a new one."""
    tid = _teacher_id_by_name(conn, name)
    if tid is not None:
        return tid
    conn.execute("INSERT INTO teachers (name) VALUES (?)", (name,))
    return conn.execute(
        "SELECT id FROM teachers WHERE name = ?", (name,)).fetchone()[0]


def _load_teachers(conn: sqlite3.Connection, wb):
    ws  = wb["Teacher load table with subject"]
    raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hdr = [_SUBJECT_NAME_MAP.get(h, h) for h in raw]
    subj_cols = hdr[2:]   # e.g. ["Chinese Language", "English Language", ...]

    # Build subject name_en → DAE code mapping ("Chinese Language" → "DAE101")
    name_to_code = {r[0]: r[1] for r in conn.execute(
        "SELECT name_en, code FROM subjects")}

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name or not str(name).strip():
            continue
        name    = str(name).strip()
        is_lec1 = row[0] is not None   # numbered row = primary (lec1)

        tid = _get_or_create_teacher(conn, name)

        for i, subj_name in enumerate(subj_cols):
            if not subj_name:
                continue
            val = row[i + 2] if (i + 2) < len(row) else None
            if not isinstance(val, (int, float)) or val <= 0:
                continue
            # Resolve header name to subject code ("Chinese Language" → "DAE101")
            subj_code = name_to_code.get(subj_name)
            if not subj_code:
                continue
            capped = min(int(val), 20)
            # All teachers across both sections contribute to lec1_quota
            # (both sections can teach as primary; quota is the total across both)
            conn.execute("""
                INSERT INTO teacher_subjects (teacher_id, subject_code, lec1_quota, backup_quota)
                VALUES (?,?,?,0)
                ON CONFLICT(teacher_id, subject_code) DO UPDATE SET
                    lec1_quota = lec1_quota + ?
            """, (tid, subj_code, capped, capped))


def _load_availability(conn: sqlite3.Connection, wb):
    """Read Teacher Availability sheet (optional). Mark unavailable slots."""
    if "Teacher Availability" not in wb.sheetnames:
        return 0   # sheet doesn't exist → all available

    ws   = wb["Teacher Availability"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return 0

    # Header row: "Teacher", "Mon 0900", "Mon 1100", ...
    header = rows[0]
    slot_cols = []   # list of (col_index, day, start_time)
    for i, h in enumerate(header):
        if i == 0 or not h:
            continue
        parts = str(h).strip().split()
        if len(parts) == 2:
            day_abbr, time = parts
            day_map = {"Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday",
                       "Thu": "Thursday", "Fri": "Friday"}
            day = day_map.get(day_abbr, day_abbr)
            slot_cols.append((i, day, time))

    count = 0
    for row in rows[1:]:
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        tid = _teacher_id_by_name(conn, name)
        if tid is None:
            continue
        for col_idx, day, start_time in slot_cols:
            val = row[col_idx] if col_idx < len(row) else None
            if val and str(val).strip().upper() == "N":
                conn.execute(
                    "INSERT OR IGNORE INTO teacher_unavailability VALUES (?,?,?)",
                    (tid, day, start_time))
                count += 1
    return count


def _load_existing_schedule(conn: sqlite3.Connection, wb):
    """Load Day/Time/Room from Class list answer (Assumption A: these are fixed)."""
    ws = wb["Class list answer"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[1]
        day  = row[8]
        if not code or not day:
            continue
        time1 = str(row[9]).strip() if row[9] else None
        time2 = str(row[10]).strip() if row[10] else None
        venue = _resolve_venue(row[11]) if row[11] else None

        code_str  = str(code).strip()
        group_code = _extract_group(code_str)
        conn.execute("""
            INSERT OR REPLACE INTO schedule
                (class_code, group_code, day, time1, time2, room_code)
            VALUES (?,?,?,?,?,?)
        """, (code_str, group_code, str(day).strip(), time1, time2, venue))
        count += 1
    return count


def _load_centre_preferences(conn: sqlite3.Connection, wb) -> int:
    """
    S1 extension: load teacher centre preferences from optional sheet
    'Teacher Centre Preference'.  Format: Teacher | Centre | Type (avoid/prefer)
    Type defaults to 'avoid' if omitted.
    """
    if "Teacher Centre Preference" not in wb.sheetnames:
        return 0
    ws = wb["Teacher Centre Preference"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        name   = str(row[0]).strip()
        centre = str(row[1]).strip().upper()
        ptype  = str(row[2]).strip().lower() if len(row) > 2 and row[2] else "avoid"
        tid = _teacher_id_by_name(conn, name)
        if tid is None:
            continue
        conn.execute(
            "INSERT OR IGNORE INTO teacher_centre_preference VALUES (?,?,?)",
            (tid, centre, ptype))
        count += 1
    return count


def _load_net_teachers(conn: sqlite3.Connection, wb) -> int:
    """
    Optional sheet 'Net Teachers': one teacher name per row (col A, from row 2).
    Marks matching teachers as is_net=1 in the teachers table.
    Sheet absence is silently ignored (returns 0).
    """
    if "Net Teachers" not in wb.sheetnames:
        return 0
    ws = wb["Net Teachers"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        tid = _teacher_id_by_name(conn, name)
        if tid is None:
            continue
        conn.execute("UPDATE teachers SET is_net=1 WHERE id=?", (tid,))
        count += 1
    return count


def build_db(wb) -> sqlite3.Connection:
    """
    Build an in-memory SQLite database from an openpyxl workbook.
    Returns the populated connection.
    """
    conn = _new_db()
    _load_subjects(conn, wb)
    _load_rooms(conn, wb)
    _load_classes(conn, wb)
    _load_teachers(conn, wb)
    n_unavail  = _load_availability(conn, wb)
    _load_centre_preferences(conn, wb)
    _load_net_teachers(conn, wb)
    n_schedule = _load_existing_schedule(conn, wb)
    conn.commit()
    return conn, n_schedule, n_unavail


# ─── Phase 2: Assign group rooms ──────────────────────────────────────────────

def assign_group_rooms(conn: sqlite3.Connection):
    """
    Assign one room per class group (same-subject same-room rule).
    Chooses the largest available room at the group's centre.
    """
    groups = conn.execute("SELECT code, centre FROM class_groups").fetchall()
    for g in groups:
        group_code, centre = g["code"], g["centre"]
        # Max students across all classes in this group
        max_students = conn.execute("""
            SELECT MAX(student_count) FROM classes WHERE group_code = ?
        """, (group_code,)).fetchone()[0] or 0

        # Largest unassigned room at the correct centre with enough capacity
        # Exclude rooms already claimed by other groups (same centre, different group)
        room = conn.execute("""
            SELECT code FROM rooms
            WHERE centre = ? AND capacity >= ?
              AND code NOT IN (SELECT room_code FROM group_rooms)
            ORDER BY capacity DESC
            LIMIT 1
        """, (centre, max_students)).fetchone()

        if room:
            conn.execute(
                "INSERT OR REPLACE INTO group_rooms VALUES (?,?)",
                (group_code, room["code"]))

    conn.commit()


# ─── Phase 2b: v4 auto-assign Day / Time / Room ───────────────────────────────

def _group_to_room_centre(group_code: str) -> str:
    """'CS1' → 'CSW',  'TW2' → 'TW',  etc."""
    prefix = re.sub(r"\d+$", "", group_code)
    return _GROUP_CENTRE_ALIAS.get(prefix, prefix)


def _allowed_centres_for_group(group_code: str) -> list:
    """Return allowed room centres for a group (primary first).
    With _TS_TM_ENABLED, TS groups may also use TM rooms as fallback.
    """
    primary = _group_to_room_centre(group_code)
    if _TS_TM_ENABLED and primary == "TS":
        return ["TS", "TM"]
    return [primary]


def _pick_room(conn: sqlite3.Connection, group_code: str, student_count: int) -> tuple:
    """
    Return (room_code, centre_used) for a class group.
    centre_used differs from home centre when J8 cross-centre fallback fires.
    Returns (None, None) if no room found anywhere within travel limit.

    Search order:
    1. Exact match from group code (TW2 → TW - TW2, CS1 → CSW - C1)
    2. Largest room at allowed centres (home + TM for TS when J1 enabled)
    3. J8 assumption: nearest centre within _H8_MAX_TRAVEL_MIN minutes
    """
    centre    = _group_to_room_centre(group_code)
    num_match = re.search(r"\d+$", group_code)

    if num_match:
        num = num_match.group()
        for candidate in (f"{centre} - {centre}{num}",
                          f"{centre} - C{num}"):   # CSW - C1 style
            # Only use the group's home room if it actually fits the class;
            # otherwise fall through to a capacity-aware room search.
            row = conn.execute(
                "SELECT code FROM rooms WHERE code = ? AND capacity >= ?",
                (candidate, student_count)).fetchone()
            if row:
                return row[0], centre

    # Fallback: largest room at each allowed centre (includes TM for TS when J1 enabled)
    for allowed_centre in _allowed_centres_for_group(group_code):
        row = conn.execute("""
            SELECT code FROM rooms
            WHERE centre = ? AND capacity >= ?
            ORDER BY capacity ASC LIMIT 1
        """, (allowed_centre, student_count)).fetchone()
        if row:
            return row[0], allowed_centre

    # J8 assumption: try nearest centres in defined proximity order
    allowed = _allowed_centres_for_group(group_code)
    for fb_centre in _NEAREST_CENTRES.get(centre, []):
        row = conn.execute("""
            SELECT code FROM rooms
            WHERE centre = ? AND capacity >= ?
            ORDER BY capacity ASC LIMIT 1
        """, (fb_centre, student_count)).fetchone()
        if fb_centre not in allowed and row:
            return row[0], fb_centre

    return None, None  # no room anywhere within travel limit


def _room_free(conn: sqlite3.Connection, room: str, day: str, slot: str) -> bool:
    row = conn.execute("""
        SELECT 1 FROM schedule
        WHERE room_code = ? AND day = ? AND (time1 = ? OR time2 = ?)
        LIMIT 1
    """, (room, day, slot, slot)).fetchone()
    return row is None


def _build_teacher_capacity(conn: sqlite3.Connection) -> dict:
    """
    Pre-compute max concurrent classes per (subject, day, start_time).
    = number of teachers who can teach that subject and are not unavailable.
    Covers both standard and TKO time slots.
    """
    cap = {}
    subjects = [r[0] for r in conn.execute(
        "SELECT DISTINCT subject_code FROM teacher_subjects WHERE lec1_quota > 0")]
    all_blocks = _TIME_BLOCKS + _TKO_TIME_BLOCKS
    all_slots  = _SINGLE_SLOTS + _TKO_SINGLE_SLOTS
    for subj in subjects:
        for day in _DAY_PRIORITY:
            for slot1, _ in all_blocks:
                start = slot1.split(" - ")[0]
                if (subj, day, start) in cap:
                    continue
                n = conn.execute("""
                    SELECT COUNT(*) FROM teacher_subjects ts
                    WHERE ts.subject_code = ? AND ts.lec1_quota > 0
                      AND ts.teacher_id NOT IN (
                          SELECT teacher_id FROM teacher_unavailability
                          WHERE day = ? AND start_time = ?
                      )
                """, (subj, day, start)).fetchone()[0]
                cap[(subj, day, start)] = n
            for slot in all_slots:
                start = slot.split(" - ")[0]
                if (subj, day, start) not in cap:
                    n = conn.execute("""
                        SELECT COUNT(*) FROM teacher_subjects ts
                        WHERE ts.subject_code = ? AND ts.lec1_quota > 0
                          AND ts.teacher_id NOT IN (
                              SELECT teacher_id FROM teacher_unavailability
                              WHERE day = ? AND start_time = ?
                          )
                    """, (subj, day, start)).fetchone()[0]
                    cap[(subj, day, start)] = n
    return cap


def _has_tiandi_violation(group_slots: list, new_slot: str) -> bool:
    """
    S2 (students only): Returns True if adding new_slot creates a gap >2 hrs
    within the same half-day (AM or PM). Lunch break is not counted.
    This function is only called for student class-group slots, never for teachers.
    """
    new_half  = _SLOT_HALF_DAY.get(new_slot)
    new_start = _SLOT_START_HOUR.get(new_slot)
    if new_half is None or new_start is None:
        return False
    for ex_slot in group_slots:
        ex_half  = _SLOT_HALF_DAY.get(ex_slot)
        ex_start = _SLOT_START_HOUR.get(ex_slot)
        if ex_half != new_half or ex_start is None:
            continue
        if abs(new_start - ex_start) > 2:
            return True
    return False


def auto_assign_schedule(conn: sqlite3.Connection,
                          cc_only: bool = False) -> tuple:
    """
    v4/v6: auto-assign Day/Time/Room for every class.

    Constraints enforced:
    - H6: TKO allows any slot with start ≥ 10:00
    - H9: A class group cannot appear at two different centres on the same day
          (exception: CSW↔SSP and TW→CSW/SSP allowed per _H9_ALLOWED_CROSS_CENTRE)
    - S2: No gap >2 hrs within same AM or PM half-day — students only
    - S3: Day rotation so same-subject classes spread across Mon-Fri

    cc_only=True  → only process classes that have a cc_group set (Phase B)
    cc_only=False → only process classes with cc_group IS NULL (Phase A, default)

    Returns (scheduled_count, unscheduled_list, h3_exceptions_list).
    unscheduled_list entries: {"code": ..., "reason_code": ..., "reason": ..., "category": ...}
    h3_exceptions_list entries: {"code": ..., "from_centre": ..., "to_centre": ..., "reason_code": ...}
    """
    from collections import defaultdict

    cc_filter = "IS NOT NULL" if cc_only else "IS NULL"
    # MRV ordering: most-constrained classes first.
    # TKO first (limited time windows), then 4-hr blocks (need two consecutive free slots),
    # then by student count desc (larger classes harder to fit), then code for stability.
    classes = conn.execute(f"""
        SELECT c.code, c.group_code, c.student_count, c.subject_code,
               sub.loading_hrs, cg.centre
        FROM classes c
        JOIN subjects sub ON sub.code = c.subject_code
        JOIN class_groups cg ON cg.code = c.group_code
        WHERE c.cc_group {cc_filter}
        ORDER BY
            CASE WHEN cg.centre = 'TKO' THEN 0 ELSE 1 END,
            CASE WHEN sub.loading_hrs >= 4 THEN 0 ELSE 1 END,
            c.student_count DESC,
            c.code ASC
    """).fetchall()

    teacher_cap      = _build_teacher_capacity(conn)
    slot_used        = defaultdict(int)   # (subject, day, start) -> count
    subj_day_offset  = defaultdict(int)   # subject -> next day rotation index
    group_day_centre = {}                 # (group, day) -> centre  [H9]
    group_day_slots  = defaultdict(list)  # (group, day) -> [slot, ...]  [S2]

    count         = 0
    unscheduled   = []   # {"code": ..., "reason_code": ..., "reason": ..., "category": ...}
    h3_exceptions = []   # H3 exceptions from J1/J8 assumptions

    for cls in classes:
        code     = cls["code"]
        group    = cls["group_code"]
        students = cls["student_count"]
        loading  = cls["loading_hrs"] or 4
        subj     = cls["subject_code"]

        room, room_centre_assigned = _pick_room(conn, group, students)
        if not room:
            centre = cls["centre"] if "centre" in cls.keys() else group
            unscheduled.append({
                "code":        code,
                "reason_code": "NO_ROOM_CAPACITY",
                "reason":      f"no room at {centre} with capacity >= {students}",
                "category":    "DATA_PROBLEM",
            })
            continue

        # Track H3 exceptions when J1/J8 fallback assigned a different centre
        home_centre = cls["centre"] if "centre" in cls.keys() else _group_to_room_centre(group)
        if room_centre_assigned != home_centre:
            rc = "H3_TS_TM" if home_centre == "TS" else "H3_FALLBACK"
            h3_exceptions.append({
                "code":        code,
                "from_centre": home_centre,
                "to_centre":   room_centre_assigned,
                "reason_code": rc,
            })

        # Derive room centre for H9/H6 checks (authoritative from DB)
        room_centre_row = conn.execute(
            "SELECT centre FROM rooms WHERE code=?", (room,)).fetchone()
        room_centre = room_centre_row[0] if room_centre_row else None

        # H6: TKO uses different time blocks
        is_tko       = (room_centre == "TKO")
        time_blocks  = _TKO_TIME_BLOCKS  if is_tko else _TIME_BLOCKS
        single_slots = _TKO_SINGLE_SLOTS if is_tko else _SINGLE_SLOTS

        offset    = subj_day_offset[subj]
        day_order = _DAY_PRIORITY[offset:] + _DAY_PRIORITY[:offset]
        need_two    = loading >= 4
        assigned    = False
        slot_blocks = {"SLOT_H9": 0, "SLOT_ROOM": 0, "SLOT_H6": 0, "SLOT_TEACHER": 0}

        for day in day_order:
            # H9: skip day if group is already at a different centre
            # Exception: allowed cross-centre pairs (CSW↔SSP, TW→CSW/SSP)
            committed = group_day_centre.get((group, day))
            if committed and room_centre and committed != room_centre:
                if frozenset({committed, room_centre}) not in _H9_ALLOWED_CROSS_CENTRE:
                    slot_blocks["SLOT_H9"] += 1
                    continue

            existing_slots = group_day_slots[(group, day)]

            if need_two:
                for slot1, slot2 in time_blocks:
                    start1 = slot1.split(" - ")[0]
                    start2 = slot2.split(" - ")[0]
                    cap = teacher_cap.get((subj, day, start1), 0)
                    if slot_used[(subj, day, start1)] >= cap:
                        slot_blocks["SLOT_TEACHER"] += 1
                        continue
                    # S2: 4-hr block occupies both AM or both PM slots consecutively
                    if (_has_tiandi_violation(existing_slots, slot1) or
                            _has_tiandi_violation(existing_slots, slot2)):
                        continue
                    if not (_room_free(conn, room, day, slot1) and
                            _room_free(conn, room, day, slot2)):
                        slot_blocks["SLOT_ROOM"] += 1
                        continue
                    conn.execute("""
                        INSERT OR REPLACE INTO schedule
                            (class_code, group_code, day, time1, time2, room_code)
                        VALUES (?,?,?,?,?,?)
                    """, (code, group, day, slot1, slot2, room))
                    slot_used[(subj, day, start1)] += 1
                    slot_used[(subj, day, start2)] += 1
                    group_day_centre[(group, day)] = room_centre
                    group_day_slots[(group, day)].extend([slot1, slot2])
                    count  += 1
                    assigned = True
                    break
            else:
                for slot in single_slots:
                    start = slot.split(" - ")[0]
                    cap = teacher_cap.get((subj, day, start), 0)
                    if slot_used[(subj, day, start)] >= cap:
                        slot_blocks["SLOT_TEACHER"] += 1
                        continue
                    # S2: check adding this slot doesn't create a gap
                    if _has_tiandi_violation(existing_slots, slot):
                        continue
                    if not _room_free(conn, room, day, slot):
                        slot_blocks["SLOT_ROOM"] += 1
                        continue
                    conn.execute("""
                        INSERT OR REPLACE INTO schedule
                            (class_code, group_code, day, time1, time2, room_code)
                        VALUES (?,?,?,?,NULL,?)
                    """, (code, group, day, slot, room))
                    slot_used[(subj, day, start)] += 1
                    group_day_centre[(group, day)] = room_centre
                    group_day_slots[(group, day)].append(slot)
                    count  += 1
                    assigned = True
                    break
            if assigned:
                break

        if assigned:
            subj_day_offset[subj] = (offset + 1) % len(_DAY_PRIORITY)
        else:
            # Determine primary blocking reason (priority: H9 > SLOT_ROOM > SLOT_TEACHER)
            # Note: TKO H6 (start ≥10:00) is pre-filtered via _TKO_TIME_BLOCKS — not an explicit block
            days_tried = len(day_order)
            if slot_blocks["SLOT_H9"] > 0 and slot_blocks["SLOT_H9"] >= days_tried:
                rc, cat = "SLOT_H9", "ALGORITHM_ISSUE"
            elif is_tko and (slot_blocks["SLOT_ROOM"] + slot_blocks["SLOT_TEACHER"]) == 0:
                rc, cat = "SLOT_H6", "ALGORITHM_ISSUE"   # TKO with no valid slots at all
            elif slot_blocks["SLOT_ROOM"] >= slot_blocks["SLOT_TEACHER"]:
                rc, cat = "SLOT_ROOM", "ALGORITHM_ISSUE"
            else:
                rc, cat = "SLOT_TEACHER", "ALGORITHM_ISSUE"
            unscheduled.append({
                "code":        code,
                "reason_code": rc,
                "reason":      f"no free slot: {rc} (H9={slot_blocks['SLOT_H9']}, room={slot_blocks['SLOT_ROOM']}, teacher={slot_blocks['SLOT_TEACHER']})",
                "category":    cat,
            })

    conn.commit()
    return count, unscheduled, h3_exceptions


# ─── Phase 2c: CC Combine scheduling (L1 / L2 / H5 / S4) ────────────────────

def _select_cc_centre(conn: sqlite3.Connection, class_codes: list) -> list:
    """
    S4: Return centres ordered by descending student count across the CC group.
    Tie-break: prefer centres closer to SSP/CSW (higher _CENTRE_RANK).
    """
    centre_students: dict = {}
    for code in class_codes:
        row = conn.execute("""
            SELECT cg.centre, c.student_count
            FROM classes c
            JOIN class_groups cg ON cg.code = c.group_code
            WHERE c.code = ?
        """, (code,)).fetchone()
        if row:
            c, s = row[0], row[1] or 0
            centre_students[c] = centre_students.get(c, 0) + s
    return sorted(centre_students.keys(),
                  key=lambda c: (-centre_students[c], -_CENTRE_RANK.get(c, 0)))


def _find_cc_days(conn: sqlite3.Connection, class_codes: list,
                  centre: str) -> list:
    """
    L1: Return ALL days (in priority order) where NONE of the CC group's
    students already have a Core class at a different centre. Returning every
    valid day lets the caller try the next one when a day's rooms are full,
    instead of aborting the whole group on the first candidate.
    """
    valid = []
    for day in _DAY_PRIORITY:
        conflict = False
        for code in class_codes:
            group = _extract_group(code)
            # Check if this group has any scheduled class at a *different* centre
            # on this day (Core classes are already scheduled at Phase A)
            row = conn.execute("""
                SELECT r.centre FROM schedule s
                JOIN rooms r ON r.code = s.room_code
                WHERE s.group_code = ? AND s.day = ?
                  AND r.centre != ?
                LIMIT 1
            """, (group, day, centre)).fetchone()
            if row:
                conflict = True
                break
        if not conflict:
            valid.append(day)
    return valid


def cc_assign_schedule(conn: sqlite3.Connection) -> tuple:
    """
    Phase B: schedule CC Combine classes (those with cc_group set).

    For each CC group:
      H5: validate all classes share the same subject_code and language.
      S4: pick centre with most students (tie-break by _CENTRE_RANK).
      L1: find a day where no student has a Core class at another centre.
      L2: if L1 fails, try next-best centre.
      → ERROR if all centres fail.

    CC classes do NOT need to follow H3 (campus restriction).
    Returns (assigned_count, error_groups).
    """
    groups = conn.execute("""
        SELECT cc_group, subject_code, language,
               GROUP_CONCAT(code) AS codes
        FROM classes
        WHERE cc_group IS NOT NULL
        GROUP BY cc_group, subject_code, language
    """).fetchall()

    assigned = 0
    errors   = []

    for grp in groups:
        cc_group   = grp["cc_group"]
        subj       = grp["subject_code"]
        lang       = grp["language"]
        codes      = grp["codes"].split(",")

        # H5: all codes must share same subject (already guaranteed by GROUP BY)
        # but log if language is mixed
        lang_check = conn.execute("""
            SELECT DISTINCT language FROM classes WHERE cc_group = ?
        """, (cc_group,)).fetchall()
        if len(lang_check) > 1:
            errors.append(f"{cc_group}: mixed languages — H5 violated, skipped")
            continue

        # S4 / L1 / L2: try each centre in preference order
        centre_order = _select_cc_centre(conn, codes)

        # CC distance filter: remove centres too far from any group's home (pending J2)
        if _CC_MAX_TRAVEL_MIN is not None:
            home_centres = set()
            for code in codes:
                row = conn.execute("""
                    SELECT cg.centre FROM classes c
                    JOIN class_groups cg ON cg.code = c.group_code
                    WHERE c.code = ?
                """, (code,)).fetchone()
                if row:
                    home_centres.add(row[0])
            centre_order = [
                c for c in centre_order
                if all(_travel_mins(home, c) <= _CC_MAX_TRAVEL_MIN for home in home_centres)
            ]

        scheduled    = False

        for centre in centre_order:
            valid_days = _find_cc_days(conn, codes, centre)
            if not valid_days:
                continue

            # Find a suitable room at this centre
            total_students = sum(
                (conn.execute("SELECT student_count FROM classes WHERE code=?",
                              (c,)).fetchone() or [0])[0]
                for c in codes
            )
            rooms = conn.execute("""
                SELECT code FROM rooms
                WHERE centre = ? AND capacity >= ?
                ORDER BY capacity ASC
            """, (centre, total_students)).fetchall()
            if not rooms:
                continue

            loading   = conn.execute(
                "SELECT loading_hrs FROM subjects WHERE code=?",
                (subj,)).fetchone()
            loading_hrs = (loading[0] if loading else 4) or 4
            need_two    = loading_hrs >= 4
            time_blocks = _TKO_TIME_BLOCKS if centre == "TKO" else _TIME_BLOCKS
            single_slots = _TKO_SINGLE_SLOTS if centre == "TKO" else _SINGLE_SLOTS

            # Try each valid day; within a day try every room with enough capacity.
            # A single full room/day must not abort the whole group.
            for day in valid_days:
                room_code = slot1 = slot2 = None
                for rm in rooms:
                    rc = rm[0]
                    if need_two:
                        for s1, s2 in time_blocks:
                            if (_room_free(conn, rc, day, s1) and
                                    _room_free(conn, rc, day, s2)):
                                room_code, slot1, slot2 = rc, s1, s2
                                break
                    else:
                        for s in single_slots:
                            if _room_free(conn, rc, day, s):
                                room_code, slot1 = rc, s
                                break
                    if slot1:
                        break

                if not slot1:
                    continue

                # Write one schedule row per CC class code
                for code in codes:
                    group = _extract_group(code)
                    conn.execute("""
                        INSERT OR REPLACE INTO schedule
                            (class_code, group_code, day, time1, time2, room_code)
                        VALUES (?,?,?,?,?,?)
                    """, (code, group, day, slot1, slot2, room_code))
                    assigned += 1

                scheduled = True
                break

            if scheduled:
                break

        if not scheduled:
            errors.append(f"{cc_group}: no valid day/centre found (L2 exhausted) — ERROR")

    conn.commit()
    return assigned, errors


# ─── Phase 3: Assign teachers ─────────────────────────────────────────────────

def assign_teachers(conn: sqlite3.Connection):
    """
    Assign Lec1 (primary) and Lec2/3 (backups) to every scheduled class.

    Four-pass Lec1 search:
      Pass 1: quota + avail + H4 cap + H8 centre (fully strict)
      Pass 2: quota relaxed, H4 + H8 still enforced
      Pass 3: quota relaxed + H4 cap relaxed (→ H4_OVERLOAD warning), H8 still enforced
      Pass 4: quota relaxed + H4 relaxed + avail ignored (→ S1 warning), H8 still enforced

    Returns (unassigned_list, warnings_list, unassigned_detail_list).
    unassigned_list: list[str] — class codes (backward compat)
    unassigned_detail_list: list[dict] — {code, subj, centre, students, reason_code, reason, category}
    warnings_list entries have keys: class, teacher, day, reason.
    """
    classes = conn.execute("""
        SELECT s.class_code, s.room_code, c.subject_code, s.day, s.time1, s.time2
        FROM schedule s
        JOIN classes c ON c.code = s.class_code
        WHERE s.day IS NOT NULL
        ORDER BY c.student_count DESC
    """).fetchall()

    unassigned        = []   # list[str] — backward compat
    warnings          = []
    unassigned_detail = []   # list[dict] — with reason codes

    for cls in classes:
        code      = cls["class_code"]
        subj      = cls["subject_code"]
        day       = cls["day"]
        time1     = cls["time1"]
        time2     = cls["time2"]
        room_code = cls["room_code"]

        # Derive room centre for H8
        centre_row = conn.execute(
            "SELECT centre FROM rooms WHERE code=?", (room_code,)
        ).fetchone() if room_code else None
        current_centre = centre_row[0] if centre_row else None

        times  = [t for t in [time1, time2] if t]
        starts = [t.split(" - ")[0].strip() for t in times]

        warn_reason = None

        # Pass 1: fully strict (quota + avail + H4 cap + H8 + centre pref)
        lec1 = _find_teacher(conn, subj, day, starts, exclude=[],
                             use_quota=True, current_centre=current_centre)
        # Pass 2: relax subject quota
        if not lec1:
            lec1 = _find_teacher(conn, subj, day, starts, exclude=[],
                                 use_quota=False, current_centre=current_centre)
        # Pass 3: ignore centre preference (S1b soft warning)
        if not lec1:
            lec1 = _find_teacher(conn, subj, day, starts, exclude=[],
                                 use_quota=False, current_centre=current_centre,
                                 ignore_centre_pref=True)
            if lec1:
                warn_reason = "S1b: teacher assigned to non-preferred centre"
        # Pass 4: relax H4 soft cap (overflow warning)
        if not lec1:
            lec1 = _find_teacher(conn, subj, day, starts, exclude=[],
                                 use_quota=False, current_centre=current_centre,
                                 ignore_h4_cap=True, ignore_centre_pref=True)
            if lec1:
                warn_reason = "H4 (soft): teacher exceeds preferred weekly loading of 6 sessions"
        # Pass 5: ignore unavailability (S1 forced, last resort)
        if not lec1:
            lec1 = _find_teacher(conn, subj, day, starts, exclude=[],
                                 use_quota=False, current_centre=current_centre,
                                 ignore_h4_cap=True, ignore_centre_pref=True,
                                 ignore_unavail=True)
            if lec1:
                warn_reason = "S1: teacher assigned to unavailable slot (loading also exceeded)"

        if not lec1:
            # Diagnostic: determine why all 5 passes failed
            qual_count = conn.execute(
                "SELECT COUNT(*) FROM teacher_subjects WHERE subject_code = ?",
                (subj,)).fetchone()[0]
            if qual_count == 0:
                t_rc, t_cat = "TEACHER_NONE", "DATA_PROBLEM"
            else:
                t_rc, t_cat = "TEACHER_CAPACITY", "DATA_PROBLEM"

            # Enrich with centre/student info for issues sheet
            centre_info = conn.execute("""
                SELECT cg.centre, c.student_count
                FROM classes c
                JOIN class_groups cg ON cg.code = c.group_code
                WHERE c.code = ?
            """, (code,)).fetchone()
            t_centre   = centre_info["centre"]        if centre_info else ""
            t_students = centre_info["student_count"] if centre_info else 0

            unassigned.append(code)
            unassigned_detail.append({
                "code":        code,
                "subj":        subj,
                "centre":      t_centre,
                "students":    t_students,
                "reason_code": t_rc,
                "reason":      f"teacher assignment failed: {t_rc} for subject {subj}",
                "category":    t_cat,
            })
        else:
            teacher_name = conn.execute(
                "SELECT name FROM teachers WHERE id=?", (lec1,)).fetchone()
            t_name = teacher_name[0] if teacher_name else str(lec1)
            if warn_reason:
                warnings.append({
                    "class":   code,
                    "teacher": t_name,
                    "day":     day,
                    "reason":  warn_reason,
                })
            # H8 soft warning: teacher crosses centres but travel time is acceptable
            if current_centre:
                committed = _get_teacher_day_centre(conn, lec1, day)
                if committed and committed != current_centre:
                    mins = _travel_mins(committed, current_centre)
                    warnings.append({
                        "class":   code,
                        "teacher": t_name,
                        "day":     day,
                        "reason":  f"H8 (soft): cross-centre {committed}↔{current_centre} ({mins} min travel)",
                    })

        lec2 = _find_teacher(conn, subj, day, starts, exclude=[lec1],
                             use_quota=False, ignore_h4_cap=True,
                             current_centre=current_centre) if lec1 else None
        lec3 = _find_teacher(conn, subj, day, starts, exclude=[lec1, lec2],
                             use_quota=False,
                             current_centre=current_centre) if lec2 else None

        conn.execute("""
            UPDATE schedule SET teacher1_id=?, teacher2_id=?, teacher3_id=?
            WHERE class_code=?
        """, (lec1, lec2, lec3, code))

        if lec1:
            conn.execute("""
                UPDATE teacher_subjects SET lec1_quota = MAX(0, lec1_quota - 1)
                WHERE teacher_id = ? AND subject_code = ?
            """, (lec1, subj))

    conn.commit()
    return unassigned, warnings, unassigned_detail


def _get_teacher_day_centre(conn, teacher_id: int, day: str):
    """Return the centre a teacher is already committed to on a given day, or None."""
    row = conn.execute("""
        SELECT r.centre FROM schedule s
        JOIN rooms r ON r.code = s.room_code
        WHERE s.teacher1_id = ? AND s.day = ?
        LIMIT 1
    """, (teacher_id, day)).fetchone()
    return row[0] if row else None


def _find_teacher(conn, subject_code, day, start_times, exclude, use_quota,
                  current_centre=None, ignore_unavail=False, ignore_h4_cap=False,
                  ignore_centre_pref=False):
    """
    Find the best available teacher for a subject at given day/start_times.

    Parameters:
      use_quota:           only consider teachers with lec1_quota > 0 (H7)
      current_centre:      exclude teachers already at a different centre today (H8)
      ignore_unavail:      skip unavailability filter (S1 forced fallback)
      ignore_h4_cap:       skip weekly session cap filter (H4 soft cap fallback)
      ignore_centre_pref:  skip centre preference filter (S1b fallback)

    Returns (teacher_id | None).
    """
    placeholders  = ",".join("?" * max(len(exclude), 1))
    exclude_safe  = exclude if exclude else [-1]
    quota_filter  = "AND ts.lec1_quota > 0" if use_quota else ""

    # H4: exclude teachers at or above weekly loading cap.
    # 1 loading = 1 class assignment (regardless of 2hr or 4hr duration).
    if ignore_h4_cap:
        h4_filter = ""
        h4_params = []
    else:
        h4_filter = """
            AND t.id NOT IN (
                SELECT teacher1_id FROM schedule
                WHERE teacher1_id IS NOT NULL
                GROUP BY teacher1_id
                HAVING COUNT(*) >= ?
            )
        """
        h4_params = [_TEACHER_WEEKLY_SESSION_CAP]

    # H8 is now a Python post-filter (travel-time based), not a SQL block.
    # Teachers at a different centre are excluded if travel time > _H8_MAX_TRAVEL_MIN.

    # S1b: centre preference filter (skipped when ignore_centre_pref=True)
    if ignore_centre_pref or not current_centre:
        cpref_filter = ""
        cpref_params = []
    else:
        cpref_filter = """
            AND t.id NOT IN (
                SELECT teacher_id FROM teacher_centre_preference
                WHERE centre = ? AND pref_type = 'avoid'
            )
        """
        cpref_params = [current_centre]

    # S1: unavailability filter (skipped when ignore_unavail=True)
    if ignore_unavail or not start_times:
        unavail_filter = ""
        unavail_params = []
    else:
        unavail_checks = " OR ".join(
            ["(tu.day = ? AND tu.start_time = ?)"] * len(start_times))
        unavail_filter = f"""
            AND t.id NOT IN (
                SELECT tu.teacher_id FROM teacher_unavailability tu
                WHERE {unavail_checks}
            )
        """
        unavail_params = [v for t in start_times for v in (day, t)]

    # Double-booking check (always enforced)
    if start_times:
        dbook_checks = " OR ".join(
            ["(s.day = ? AND (s.time1 LIKE ? OR s.time2 LIKE ?))"] * len(start_times))
        dbook_filter = f"""
            AND t.id NOT IN (
                SELECT s.teacher1_id FROM schedule s
                WHERE s.teacher1_id IS NOT NULL AND ({dbook_checks})
            )
        """
        dbook_params = [v for t in start_times for v in (day, f"%{t}%", f"%{t}%")]
    else:
        dbook_filter = ""
        dbook_params = []

    query = f"""
        SELECT t.id
        FROM teachers t
        JOIN teacher_subjects ts ON t.id = ts.teacher_id
        WHERE ts.subject_code = ?
          {quota_filter}
          AND t.id NOT IN ({placeholders})
          {h4_filter}
          {cpref_filter}
          {unavail_filter}
          {dbook_filter}
        ORDER BY ts.lec1_quota DESC
        LIMIT 20
    """
    params = ([subject_code] + exclude_safe + h4_params +
              cpref_params + unavail_params + dbook_params)
    rows = conn.execute(query, params).fetchall()

    # H8 post-filter: allow cross-centre only if travel time <= _H8_MAX_TRAVEL_MIN
    for row in rows:
        teacher_id = row[0]
        if current_centre:
            committed = _get_teacher_day_centre(conn, teacher_id, day)
            if committed and committed != current_centre:
                if _travel_mins(committed, current_centre) > _H8_MAX_TRAVEL_MIN:
                    continue  # hard block: travel too far
        return teacher_id
    return None


# ─── Phase 3b: English Net teacher weekly assignment ─────────────────────────

def _load_english_weekly_from_wb(wb) -> dict:
    """
    Read pre-assigned English teacher data from the "English Weekly" sheet.
    Returns {(class_code, term, block): teacher_name} or empty dict if sheet absent.

    Expected sheet layout:
      Row 1: headers (Class Code, T2025C wk1-5, T2025C wk6-10, T2025C wk11-15,
                       T2026A wk1-5, T2026A wk6-10, T2026A wk11-15)
      Row 2+: data rows
    """
    if "English Weekly" not in wb.sheetnames:
        return {}
    ws = wb["English Weekly"]
    result = {}
    term_block_cols = [
        ("T2025C", "wk1-5"),   # col B (index 1)
        ("T2025C", "wk6-10"),  # col C (index 2)
        ("T2025C", "wk11-15"), # col D (index 3)
        ("T2026A", "wk1-5"),   # col E (index 4)
        ("T2026A", "wk6-10"),  # col F (index 5)
        ("T2026A", "wk11-15"), # col G (index 6)
    ]
    for row in ws.iter_rows(min_row=2, values_only=True):
        class_code = row[0]
        if not class_code:
            continue
        class_code = str(class_code).strip()
        for i, (term, block) in enumerate(term_block_cols):
            teacher = row[i + 1] if len(row) > i + 1 else None
            if teacher:
                result[(class_code, term, block)] = str(teacher).strip()
    return result


def assign_english_weekly(conn: sqlite3.Connection,
                           preassigned: dict = None) -> tuple:
    """
    Use pre-assigned English (DAE102) teacher data and validate constraints.

    Priority:
      1. preassigned dict (from "English Weekly" Excel sheet)
      2. _ENG_WEEKLY_PREASSIGNED constant (hard-coded for current term)
      3. Auto-assign fallback (legacy algorithm)

    Checks:
      - Teacher availability (B44): warn if assigned teacher unavailable on class day
      - Net teacher requirement (B46): warn if ≥2 Net blocks not met per term
      - CS1-3, CS7 exempt from Net requirement (B47/B48)

    Returns (assignments_dict, warnings_list).
    assignments_dict: {(class_code, term, block): teacher_name or None}
    """
    # Resolve which assignment source to use
    source = preassigned if preassigned else _ENG_WEEKLY_PREASSIGNED
    use_preassigned = bool(source)

    # All scheduled English classes (day/time for availability check)
    eng_classes = conn.execute("""
        SELECT s.class_code, s.day, s.time1, cg.centre, c.group_code
        FROM schedule s
        JOIN classes c ON c.code = s.class_code
        JOIN class_groups cg ON cg.code = c.group_code
        WHERE c.subject_code = ? AND s.day IS NOT NULL
        ORDER BY s.day, s.time1, s.class_code
    """, (_ENG_SUBJECT_CODE,)).fetchall()

    if not eng_classes:
        return {}, []

    # Build lookup: class_code → {day, time1, centre, group}
    cls_info = {r["class_code"]: dict(r) for r in eng_classes}

    # Net teacher info for validation
    all_eng = conn.execute("""
        SELECT t.id, t.name, t.is_net
        FROM teachers t
        JOIN teacher_subjects ts ON ts.teacher_id = t.id
        WHERE ts.subject_code = ?
    """, (_ENG_SUBJECT_CODE,)).fetchall()
    net_names  = {r["name"] for r in all_eng if r["is_net"]}
    net_cores  = {_canon_core(r["name"]) for r in all_eng if r["is_net"]}
    name_to_id = {r["name"]: r["id"] for r in all_eng}
    # Honorific-tolerant lookup so "Ms. Cherry Ip" in the English sheet still
    # resolves to "Mr. Cherry Ip" in the load table (and vice versa).
    core_to_id = {_canon_core(r["name"]): r["id"] for r in all_eng}

    warnings   = []
    assignments = {}   # (class_code, term, block) → teacher_name

    if use_preassigned:
        # ── Mode 1: use pre-assigned teachers ────────────────────────────────
        # Collect all class×term pairs that appear in source
        cls_terms = set((k[0], k[1]) for k in source)

        for class_code, term in sorted(cls_terms):
            info  = cls_info.get(class_code)
            if not info:
                continue
            day   = info["day"]
            time1 = info["time1"]
            start = time1.split(" - ")[0].strip() if time1 else ""
            group = info.get("group_code", "")
            exempt = group in _ENG_NET_EXEMPT_GROUPS
            net_count = 0

            for block in _ENG_WEEK_BLOCKS:
                teacher_name = source.get((class_code, term, block))
                assignments[(class_code, term, block)] = teacher_name

                if not teacher_name:
                    continue

                # Availability check (B44): warn if teacher marked unavailable
                tid = name_to_id.get(teacher_name) or core_to_id.get(_canon_core(teacher_name))
                if tid and day and start:
                    unavail = conn.execute("""
                        SELECT 1 FROM teacher_unavailability
                        WHERE teacher_id=? AND day=? AND start_time=?
                    """, (tid, day, start)).fetchone()
                    if unavail:
                        warnings.append({
                            "class":   class_code,
                            "teacher": teacher_name,
                            "day":     day,
                            "reason":  f"English Weekly: {teacher_name} unavailable on {day} {time1} ({term} {block})",
                        })

                if teacher_name in net_names or _canon_core(teacher_name) in net_cores:
                    net_count += 1

            # Net requirement check (B46)
            if not exempt and net_count < _ENG_NET_MIN_BLOCKS:
                warnings.append({
                    "class":       class_code,
                    "code":        class_code,
                    "centre":      info.get("centre", ""),
                    "teacher":     "",
                    "day":         term,
                    "reason_code": "NET_SHORTFALL",
                    "reason": (
                        f"English Net shortfall: {net_count}/{_ENG_NET_MIN_BLOCKS} Net block in {term} "
                        f"for {class_code} (pre-assigned English Weekly). Assign a Net teacher to a block "
                        f"for this class, or accept no Net coverage."
                    ),
                })

    else:
        # ── Mode 2: auto-assign fallback (legacy algorithm) ──────────────────
        from collections import defaultdict
        net_teachers   = [r["id"] for r in all_eng if r["is_net"]]
        local_teachers = [r["id"] for r in all_eng if not r["is_net"]]
        id_to_name_map = {r["id"]: r["name"] for r in all_eng}
        net_available  = bool(net_teachers)

        # How many English classes sit on each (day, time) slot — used to explain
        # a Net shortfall (crowded slot + travel isolation, not lack of teachers).
        slot_load = {}
        for c in eng_classes:
            k = (c["day"], c["time1"])
            slot_load[k] = slot_load.get(k, 0) + 1
        net_name_list = [id_to_name_map[t] for t in net_teachers if t in id_to_name_map]

        for term in _ENG_TERMS:
            block_day_used = defaultdict(set)
            block_centre: dict = {}

            for cls in eng_classes:
                class_code = cls["class_code"]
                day        = cls["day"]
                time1      = cls["time1"]
                centre     = cls["centre"]
                group      = cls["group_code"]
                exempt     = group in _ENG_NET_EXEMPT_GROUPS
                need_net   = 0 if (exempt or not net_available) else _ENG_NET_MIN_BLOCKS
                net_given  = 0
                start      = time1.split(" - ")[0].strip() if time1 else ""

                # Teachers marked unavailable on this class's day/slot are off-limits.
                unavail_here = set()
                if day and start:
                    unavail_here = {r[0] for r in conn.execute(
                        "SELECT teacher_id FROM teacher_unavailability WHERE day=? AND start_time=?",
                        (day, start)).fetchall()}

                for block in _ENG_WEEK_BLOCKS:
                    used = block_day_used[(block, day, time1)]
                    pool = (net_teachers + local_teachers) if net_given < need_net \
                           else (local_teachers + net_teachers)
                    chosen = None
                    for tid in pool:
                        if tid in used or tid in unavail_here:
                            continue
                        committed_centre = block_centre.get((block, tid, day))
                        if committed_centre and committed_centre != centre:
                            if _travel_mins(committed_centre, centre) > _ENG_MAX_TRAVEL_MIN:
                                continue
                        chosen = tid
                        break
                    t_name = id_to_name_map.get(chosen) if chosen else None
                    assignments[(class_code, term, block)] = t_name
                    if chosen:
                        used.add(chosen)
                        block_centre[(block, chosen, day)] = centre
                        if chosen in net_teachers:
                            net_given += 1

                # Net shortfall check — report WHY (crowded slot + travel isolation)
                if not exempt and net_given < need_net:
                    n_share = slot_load.get((day, time1), 1) - 1
                    net_lbl = " / ".join(
                        n.replace("Mr. ", "").replace("Ms. ", "") for n in net_name_list
                    ) or "Net teachers"
                    warnings.append({
                        "class":       class_code,
                        "code":        class_code,
                        "centre":      centre,
                        "teacher":     "",
                        "day":         term,
                        "reason_code": "NET_SHORTFALL",
                        "reason": (
                            f"English Net shortfall: {net_given}/{_ENG_NET_MIN_BLOCKS} Net block in {term} "
                            f"— {class_code} at {centre} {day} {start} shares this slot with {n_share} other "
                            f"English class(es); no Net teacher ({net_lbl}) is both free and within "
                            f"{_ENG_MAX_TRAVEL_MIN} min travel. Reschedule to a lighter slot or accept no Net coverage."
                        ),
                    })

    return assignments, warnings


# ─── Phase 4: Collect results ─────────────────────────────────────────────────

def collect_results(conn: sqlite3.Connection) -> list:
    """Return schedule as list of dicts for write_output_wb."""
    rows = conn.execute("""
        SELECT
            s.class_code, s.day, s.time1, s.time2, s.room_code,
            t1.name AS lec1, t2.name AS lec2, t3.name AS lec3,
            c.student_count,
            c.cc_group,
            sub.name_cn,
            sub.name_en,
            sub.loading_hrs
        FROM schedule s
        LEFT JOIN teachers t1 ON t1.id = s.teacher1_id
        LEFT JOIN teachers t2 ON t2.id = s.teacher2_id
        LEFT JOIN teachers t3 ON t3.id = s.teacher3_id
        JOIN classes c ON c.code = s.class_code
        JOIN subjects sub ON sub.code = c.subject_code
        WHERE s.day IS NOT NULL
        ORDER BY s.class_code
    """).fetchall()
    return [dict(r) for r in rows]


# ─── Phase 5: Write Excel output ──────────────────────────────────────────────

_TIME_SLOTS = ["0900 - 1100", "1000 - 1200", "1100 - 1300", "1200 - 1400",
               "1400 - 1600", "1500 - 1700", "1600 - 1800", "1700 - 1900"]
_DAYS_ORDER  = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
_DAY_ABBR    = {"Monday": "Mon", "Tuesday": "Tue", "Wednesday": "Wed",
                "Thursday": "Thu", "Friday": "Fri"}
_DAY_OFFSET  = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3, "Friday": 4}

# HK Public Holidays 2025–2026 (ISO format YYYY-MM-DD)
_HK_HOLIDAYS = {
    # 2025
    "2025-01-01",                                # New Year's Day
    "2025-01-29", "2025-01-30", "2025-01-31",   # Lunar New Year (Snake)
    "2025-04-04",                                # Ching Ming Festival
    "2025-04-18", "2025-04-19", "2025-04-21",   # Good Friday / Easter
    "2025-05-01",                                # Labour Day
    "2025-05-05",                                # Buddha's Birthday
    "2025-06-02",                                # Tuen Ng Festival
    "2025-07-01",                                # HKSAR Establishment Day
    "2025-10-01",                                # National Day
    "2025-10-07",                                # Day after Mid-Autumn Festival
    "2025-10-29",                                # Chung Yeung Festival
    "2025-12-25", "2025-12-26",                 # Christmas
    # 2026
    "2026-01-01",                                # New Year's Day
    "2026-02-17", "2026-02-18", "2026-02-19",   # Lunar New Year (Horse)
    "2026-04-03", "2026-04-04", "2026-04-06",   # Good Friday / Easter
    "2026-04-05",                                # Ching Ming Festival (substitute Apr 6 if clash)
    "2026-05-01",                                # Labour Day
    "2026-05-25",                                # Buddha's Birthday
    "2026-06-19",                                # Tuen Ng Festival
    "2026-07-01",                                # HKSAR Establishment Day
    "2026-09-25",                                # Day after Mid-Autumn Festival
    "2026-10-01",                                # National Day
    "2026-10-20",                                # Chung Yeung Festival
    "2026-12-25", "2026-12-26",                 # Christmas
}


def generate_class_dates(start_date_str: str, day_name: str, n_weeks: int = 15):
    """Return n_weeks dates for a class on day_name, starting from the week of start_date.
    start_date_str: 'YYYY-MM-DD' (should be the Monday of week 1).
    Returns list of datetime.date objects."""
    from datetime import date, timedelta
    try:
        term_start = date.fromisoformat(start_date_str)
    except (ValueError, AttributeError):
        return []
    monday = term_start - timedelta(days=term_start.weekday())
    offset = _DAY_OFFSET.get(day_name, 0)
    return [monday + timedelta(weeks=w, days=offset) for w in range(n_weeks)]


def write_output_fast(results: list, english_weekly: dict = None,
                      issues: list = None, assumptions: list = None,
                      term_dates: dict = None) -> bytes:
    """Build a new clean workbook from results — no original formatting loaded.
    Runs in < 1 s vs 47 s for the modify-in-place approach."""
    from io import BytesIO

    wb = openpyxl.Workbook()

    # ── Sheet 1: Class Assignments ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Class Assignments"
    ws1.append(["Class Code", "Subject", "Subject (CN)", "Room", "Day",
                "Time", "Students", "Lec 1", "Lec 2", "Lec 3"])
    for r in sorted(results, key=lambda x: (x["day"] or "", x["time1"] or "",
                                             x["class_code"])):
        ws1.append([
            r["class_code"],
            r["name_en"] or "",
            r["name_cn"] or "",
            r["room_code"] or "",
            r["day"] or "",
            r["time1"] or "",
            r["student_count"] or 0,
            r["lec1"] or "",
            r["lec2"] or "",
            r["lec3"] or "",
        ])

    # ── Sheets 2-6: Daily timetable grids ────────────────────────────────────
    for day in _DAYS_ORDER:
        day_results = [r for r in results if r["day"] == day]
        if not day_results:
            continue

        rooms  = sorted(set(r["room_code"] for r in day_results if r["room_code"]))
        lookup = {(r["room_code"], r["time1"]): r
                  for r in day_results if r["room_code"] and r["time1"]}

        ws = wb.create_sheet(_DAY_ABBR[day])
        ws.append(["Room"] + _TIME_SLOTS)
        for room in rooms:
            row_data = [room]
            for slot in _TIME_SLOTS:
                entry = lookup.get((room, slot))
                if entry:
                    row_data.append(
                        f"{entry['class_code']} ({entry['student_count']})\n"
                        f"Lec1: {entry['lec1'] or '-'}"
                    )
                else:
                    row_data.append("")
            ws.append(row_data)

    # ── Sheet 7: English Weekly Teacher Assignment ───────────────────────────
    if english_weekly:
        eng_results = [r for r in results if r["name_en"] == _ENG_SUBJECT_CODE
                       or (r.get("name_en") or "").startswith("English")]
        # Fallback: match by class_code key in english_weekly
        coded = {k[0] for k in english_weekly}

        ws_eng = wb.create_sheet("English Weekly")
        # Header: Class | Day | Time | Room | T2025C wk1-5 | wk6-10 | wk11-15 | T2026A wk1-5 | wk6-10 | wk11-15
        header = ["Class Code", "Day", "Time", "Room"]
        for term in _ENG_TERMS:
            for blk in _ENG_WEEK_BLOCKS:
                header.append(f"{term} {blk}")
        ws_eng.append(header)

        eng_rows = [r for r in results if r["class_code"] in coded]
        for r in sorted(eng_rows, key=lambda x: (x["day"] or "", x["time1"] or "", x["class_code"])):
            row_data = [r["class_code"], r["day"] or "", r["time1"] or "", r["room_code"] or ""]
            for term in _ENG_TERMS:
                for blk in _ENG_WEEK_BLOCKS:
                    row_data.append(english_weekly.get((r["class_code"], term, blk)) or "")
            ws_eng.append(row_data)

    # ── Issues sheet ─────────────────────────────────────────────────────────
    if issues:
        from openpyxl.styles import Font
        ws_issues = wb.create_sheet("Issues")
        headers = ["Class Code", "Centre", "Subject", "Students",
                   "Category", "Reason Code", "Reason", "Suggested Action"]
        ws_issues.append(headers)
        for cell in ws_issues[1]:
            cell.font = Font(bold=True)
        cat_order = {"DATA_PROBLEM": 0, "ALGORITHM_ISSUE": 1, "ASSUMPTION": 2}
        sorted_issues = sorted(
            issues, key=lambda x: cat_order.get(x.get("category", "DATA_PROBLEM"), 3))
        for item in sorted_issues:
            ws_issues.append([
                item.get("code", ""),
                item.get("centre", ""),
                item.get("subj", ""),
                item.get("students", 0),
                item.get("category", ""),
                item.get("reason_code", ""),
                item.get("reason", ""),
                item.get("suggested", ""),
            ])
        if assumptions:
            ws_issues.append([])
            ws_issues.append(["--- ASSUMPTIONS (pending Jo confirmation) ---"])
            ws_issues.append(["Question", "Assumption", "Count", "Affected Classes"])
            for a in assumptions:
                affected = a.get("affected", [])
                ws_issues.append([
                    a.get("question", ""),
                    a.get("assumption", ""),
                    len(affected),
                    ", ".join(affected[:20]),
                ])

    # ── Term Date Sheets ──────────────────────────────────────────────────────
    if term_dates:
        from openpyxl.styles import PatternFill, Font as _Font
        from datetime import date as _date
        red_fill = PatternFill("solid", fgColor="FF4444")
        hdr_font = _Font(bold=True)

        # Build lookup: class_code → day
        day_lookup = {r["class_code"]: r["day"] for r in results if r["class_code"] and r["day"]}

        for term_code, start_date_str in sorted(term_dates.items()):
            if not start_date_str:
                continue
            ws_dt = wb.create_sheet(f"{term_code} Dates")
            hdr = ["Class Code", "Day"] + [f"Wk {i+1}" for i in range(15)]
            ws_dt.append(hdr)
            for cell in ws_dt[1]:
                cell.font = hdr_font

            seen = set()
            for r in sorted(results, key=lambda x: (x["day"] or "", x["class_code"] or "")):
                cc = r["class_code"]
                day_name = r["day"]
                if not cc or not day_name or cc in seen:
                    continue
                seen.add(cc)
                dates = generate_class_dates(start_date_str, day_name)
                row_data = [cc, day_name] + [d.strftime("%d/%m/%Y") for d in dates]
                ws_dt.append(row_data)
                row_idx = ws_dt.max_row
                for col_idx, d in enumerate(dates, start=3):
                    if d.strftime("%Y-%m-%d") in _HK_HOLIDAYS:
                        ws_dt.cell(row=row_idx, column=col_idx).fill = red_fill

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_output(input_path: str, output_path: str, results: list):
    from io import BytesIO
    data = write_output_fast(results)
    with open(output_path, "wb") as f:
        f.write(data)
    print(f"\n  Saved: {output_path}")


# ─── Stats & validation ───────────────────────────────────────────────────────

def collect_stats(conn: sqlite3.Connection, results: list,
                  unassigned: list, s1_warnings: list = None) -> dict:
    total   = len(results)
    lec1    = sum(1 for r in results if r["lec1"])
    lec2    = sum(1 for r in results if r["lec2"])
    lec3    = sum(1 for r in results if r["lec3"])
    unavail = conn.execute(
        "SELECT COUNT(*) FROM teacher_unavailability").fetchone()[0]

    # Teacher loading: sessions + hours per teacher
    t_sessions: dict = {}
    t_hours: dict    = {}
    for r in results:
        name = r.get("lec1")
        if not name:
            continue
        hrs = r.get("loading_hrs") or 4
        t_sessions[name] = t_sessions.get(name, 0) + 1
        t_hours[name]    = t_hours.get(name, 0) + hrs
    teacher_loading = sorted(
        [{"name": n, "sessions": t_sessions[n], "hours": t_hours[n]}
         for n in t_sessions],
        key=lambda x: -x["hours"]
    )

    # Build timetable grid: day -> room -> slot -> class info
    grid: dict = {}
    for r in results:
        day  = r.get("day")
        room = r.get("room_code") or ""
        if not day or not room:
            continue
        grid.setdefault(day, {}).setdefault(room, {})
        for slot in [r.get("time1"), r.get("time2")]:
            if slot:
                grid[day][room][slot] = {
                    "code":       r["class_code"],
                    "subject_cn": r["name_cn"],
                    "subject_en": r["name_en"],
                    "lec1":       r.get("lec1") or "",
                    "students":   r.get("student_count") or 0,
                }

    # All rooms sorted by centre then code — for grid display (includes empty rooms)
    all_rooms = [r["code"] for r in conn.execute(
        "SELECT code FROM rooms ORDER BY centre, code").fetchall()]

    return {
        "scheduled":       lec1,
        "total_classes":   total,
        "unassigned":      unassigned,
        "warnings":        s1_warnings or [],
        "violations":      [],
        "lec2_coverage":   lec2,
        "lec3_coverage":   lec3,
        "unavail_slots":   unavail,
        "preferred_pct":   100,
        "centre_dist":     {},
        "teacher_loading": teacher_loading,
        "timetable_grid":  grid,
        "all_rooms":       all_rooms,
        "results":         results,
    }


def validate_and_report(conn, results, unassigned):
    stats = collect_stats(conn, results, unassigned)
    sep = "=" * 62
    print(f"\n{sep}")
    print("  TIMETABLE SCHEDULER v3 — VALIDATION REPORT")
    print(sep)
    print(f"  Classes with Lec1    : {stats['scheduled']}/{stats['total_classes']}")
    print(f"  Lec2 coverage        : {stats['lec2_coverage']}/{stats['total_classes']}")
    print(f"  Lec3 coverage        : {stats['lec3_coverage']}/{stats['total_classes']}")
    print(f"  Unavailability slots : {stats['unavail_slots']}")

    if unassigned:
        print(f"\n  No eligible teacher ({len(unassigned)}):")
        for c in sorted(unassigned):
            print(f"    - {c}")

    if stats["violations"]:
        print(f"\n  Room note ({len(stats['violations'])} groups use different rooms")
        print("  across subjects — consistent with existing Class list answer data):")
        for v in stats["violations"][:5]:
            print(f"    - {v}")
        if len(stats["violations"]) > 5:
            print(f"    ... and {len(stats['violations'])-5} more")

    status = "GO" if not unassigned else "REVIEW NEEDED"
    print(f"\n  Overall: {status}")
    print(sep)


# ─── Web API entry point ──────────────────────────────────────────────────────

def run_from_bytes(excel_bytes: bytes) -> tuple:
    """
    Process an uploaded Excel file in memory.
    Returns (output_excel_bytes, stats_dict).

    Memory strategy: use read_only=True for the read phase (streaming,
    ~10x less memory), close and gc before loading the write copy.
    This keeps peak memory well under Vercel's 1024MB limit.
    """
    import gc
    from io import BytesIO

    # ── Phase 1: READ (streaming, low memory) ────────────────────────────────
    wb_read = openpyxl.load_workbook(
        BytesIO(excel_bytes), data_only=True, read_only=True)

    conn, n_sched, n_unavail = build_db(wb_read)
    wb_read.close()
    del wb_read
    gc.collect()

    if n_sched == 0:
        raise ValueError(
            "Class list answer has no scheduled data. "
            "Please upload the original Planning for Timetable.xlsx, "
            "not a previous output file.")

    unassigned, s1_warnings = assign_teachers(conn)
    results    = collect_results(conn)
    stats      = collect_stats(conn, results, unassigned, s1_warnings)
    del conn
    gc.collect()

    # ── Phase 2: WRITE (new clean workbook — no original formatting needed) ──
    output_bytes = write_output_fast(results)
    return output_bytes, stats


_CADET_DAYS = {"Monday", "Wednesday", "Friday"}
# Cadet day preference order (Mon > Wed > Fri). Kept separate from _DAY_PRIORITY
# so that restricting DAE subjects to Mon/Tue/Thu does not affect cadet placement,
# and so nothing calls _DAY_PRIORITY.index() on Wed/Fri (which are no longer in it).
_CADET_DAY_PRIORITY = ["Monday", "Wednesday", "Friday"]


def _is_cadet_class(class_code: str) -> bool:
    """Return True for cadet classes identified by single-letter group code (e.g. DAE256_E)."""
    parts = class_code.split("_")
    if len(parts) < 2:
        return False
    group = parts[-1]
    return len(group) == 1 and group.isalpha()


def phase0_schedule_cadets(conn: sqlite3.Connection) -> tuple:
    """
    Phase 0: Pre-schedule cadet classes (single-letter group code) on Mon/Wed/Fri.
    Must run before auto_assign_schedule() so their rooms are blocked first.
    Returns (scheduled_count, error_list).
    """
    cadet_classes = [
        row for row in conn.execute(
            "SELECT code, group_code, subject_code, student_count FROM classes"
        ).fetchall()
        if _is_cadet_class(row["code"])
    ]

    if not cadet_classes:
        return 0, []

    scheduled = 0
    errors: list = []

    for cls in cadet_classes:
        centre_row = conn.execute(
            "SELECT centre FROM class_groups WHERE code = ?", (cls["group_code"],)
        ).fetchone()
        if not centre_row:
            errors.append(f"{cls['code']}: no centre found for group {cls['group_code']}")
            continue

        centre = centre_row[0]
        room = conn.execute("""
            SELECT code FROM rooms
            WHERE centre = ? AND capacity >= ?
            ORDER BY capacity ASC LIMIT 1
        """, (centre, cls["student_count"] or 0)).fetchone()
        if not room:
            errors.append(f"{cls['code']}: no room at {centre} with capacity >= {cls['student_count']}")
            continue

        subj = cls["subject_code"]
        loading = conn.execute(
            "SELECT loading_hrs FROM subjects WHERE code = ?", (subj,)
        ).fetchone()
        loading_hrs = (loading[0] if loading else 4) or 4
        need_two    = loading_hrs >= 4
        time_blocks  = _TKO_TIME_BLOCKS if centre == "TKO" else _TIME_BLOCKS
        single_slots = _TKO_SINGLE_SLOTS if centre == "TKO" else _SINGLE_SLOTS

        assigned = False
        for day in _CADET_DAY_PRIORITY:
            if need_two:
                for s1, s2 in time_blocks:
                    if _room_free(conn, room[0], day, s1) and _room_free(conn, room[0], day, s2):
                        conn.execute("""
                            INSERT OR REPLACE INTO schedule
                                (class_code, group_code, day, time1, time2, room_code)
                            VALUES (?,?,?,?,?,?)
                        """, (cls["code"], cls["group_code"], day, s1, s2, room[0]))
                        scheduled += 1
                        assigned = True
                        break
            else:
                for s in single_slots:
                    if _room_free(conn, room[0], day, s):
                        conn.execute("""
                            INSERT OR REPLACE INTO schedule
                                (class_code, group_code, day, time1, time2, room_code)
                            VALUES (?,?,?,?,NULL,?)
                        """, (cls["code"], cls["group_code"], day, s, room[0]))
                        scheduled += 1
                        assigned = True
                        break
            if assigned:
                break

        if not assigned:
            errors.append(f"{cls['code']}: no free slot on Mon/Wed/Fri at {centre}")

    conn.commit()
    return scheduled, errors


def run_v4_from_bytes(excel_bytes: bytes, term_dates: dict = None) -> tuple:
    """
    v4 entry point: auto-assigns Day/Time/Room when Class list answer is empty.
    Falls back to v3 behaviour if Class list answer is already filled.
    term_dates: optional dict {"T2025C": "YYYY-MM-DD", "T2026A": "YYYY-MM-DD"}
    """
    import gc
    from io import BytesIO

    wb_read = openpyxl.load_workbook(
        BytesIO(excel_bytes), data_only=True, read_only=True)

    conn, n_sched, n_unavail = build_db(wb_read)
    # Load English Weekly pre-assignments before closing workbook
    eng_preassigned_from_wb = _load_english_weekly_from_wb(wb_read)
    wb_read.close()
    del wb_read
    gc.collect()

    unscheduled_rooms = []
    core_h3           = []
    unassigned_detail = []
    all_issues        = []
    assumptions_list  = []
    if n_sched == 0:
        # Phase 0: pre-schedule cadet classes (fixed Mon/Wed/Fri rooms)
        _cadet_count, cadet_errors = phase0_schedule_cadets(conn)
        # Phase A: schedule Core classes (no cc_group)
        _core_count, core_unscheduled, core_h3 = auto_assign_schedule(conn, cc_only=False)
        unscheduled_rooms.extend(core_unscheduled)
        # Phase B: schedule CC Combine classes (cc_group set)
        cc_assigned, cc_errors = cc_assign_schedule(conn)
    else:
        cadet_errors = []
        cc_errors    = []
    # else: Class list answer already filled → use existing schedule as-is

    unassigned, s1_warnings, unassigned_detail = assign_teachers(conn)
    # Phase 3b: English weekly — use pre-assigned data from Excel or hard-coded constant
    english_weekly, eng_warnings = assign_english_weekly(
        conn, preassigned=eng_preassigned_from_wb or None)
    results        = collect_results(conn)
    stats          = collect_stats(conn, results, unassigned, s1_warnings)
    stats["unscheduled_rooms"]    = unscheduled_rooms
    stats["english_weekly_count"] = len(english_weekly)
    # stats["issues"] and stats["assumptions"] set after enrichment block below

    all_errors = cadet_errors + cc_errors if n_sched == 0 else []
    if all_errors:
        stats.setdefault("warnings", [])
        for e in cadet_errors:
            stats["warnings"].append({"class": e, "teacher": "", "day": "", "reason": "Cadet Phase0 ERROR"})
        for e in cc_errors:
            stats["warnings"].append({"class": e, "teacher": "", "day": "", "reason": "CC Combine ERROR"})

    # Merge English weekly warnings (availability + Net shortfall)
    if eng_warnings:
        stats.setdefault("warnings", [])
        stats["warnings"].extend(eng_warnings)

    # Collect Net teacher names before conn is released
    net_names = {r["name"] for r in conn.execute(
        "SELECT name FROM teachers WHERE is_net=1").fetchall()}

    # Enrich unscheduled items with subject/centre/students (must be before del conn)
    enriched_unscheduled = []
    for item in unscheduled_rooms:
        row = conn.execute("""
            SELECT c.student_count, c.subject_code, cg.centre
            FROM classes c
            JOIN class_groups cg ON cg.code = c.group_code
            WHERE c.code = ?
        """, (item["code"],)).fetchone()
        enriched_unscheduled.append({
            **item,
            "students": row["student_count"]  if row else 0,
            "subj":     row["subject_code"]   if row else "",
            "centre":   row["centre"]          if row else "",
        })
    unscheduled_rooms = enriched_unscheduled

    # Build unified issues list with categories and suggested actions
    all_issues = []
    for item in unscheduled_rooms:
        rc = item.get("reason_code", "NO_ROOM_CAPACITY")
        all_issues.append({
            "code":        item["code"],
            "subj":        item.get("subj", ""),
            "centre":      item.get("centre", ""),
            "students":    item.get("students", 0),
            "reason_code": rc,
            "reason":      item["reason"],
            "category":    item.get("category", "DATA_PROBLEM"),
            "suggested":   _SUGGESTED_ACTIONS.get(rc, "Contact Jo"),
        })
    for item in unassigned_detail:
        rc = item.get("reason_code", "TEACHER_NONE")
        all_issues.append({
            "code":        item["code"],
            "subj":        item.get("subj", ""),
            "centre":      item.get("centre", ""),
            "students":    item.get("students", 0),
            "reason_code": rc,
            "reason":      item["reason"],
            "category":    item.get("category", "DATA_PROBLEM"),
            "suggested":   _SUGGESTED_ACTIONS.get(rc, "Contact Jo"),
        })
    for exc in core_h3:
        rc = exc.get("reason_code", "H3_FALLBACK")
        all_issues.append({
            "code":        exc["code"],
            "subj":        "",
            "centre":      exc.get("from_centre", ""),
            "students":    0,
            "reason_code": rc,
            "reason":      f"H3 exception: {exc.get('from_centre','')} → {exc.get('to_centre','')}",
            "category":    "ASSUMPTION",
            "suggested":   _SUGGESTED_ACTIONS.get(rc, "Confirm with Jo"),
        })
    # English Net shortfalls: surface in the Issues panel (deduped by class) so Jo
    # sees the constraint and can decide whether to reschedule. One row per class.
    seen_net = set()
    for w in eng_warnings:
        if w.get("reason_code") != "NET_SHORTFALL":
            continue
        code = w.get("code") or w.get("class", "")
        if code in seen_net:
            continue
        seen_net.add(code)
        all_issues.append({
            "code":        code,
            "subj":        _ENG_SUBJECT_CODE,
            "centre":      w.get("centre", ""),
            "students":    0,
            "reason_code": "NET_SHORTFALL",
            "reason":      w["reason"],
            "category":    "ALGORITHM_ISSUE",
            "suggested":   _SUGGESTED_ACTIONS.get("NET_SHORTFALL", "Contact Jo"),
        })

    # Build assumptions list with per-class centre detail (for UI, Excel, grid highlight)
    j1_entries = [e for e in core_h3 if e.get("reason_code") == "H3_TS_TM"]
    j8_entries = [e for e in core_h3 if e.get("reason_code") == "H3_FALLBACK"]
    assumptions_list = []
    if j1_entries:
        assumptions_list.append({
            "question":   "J1",
            "assumption": "TS→TM cross-centre enabled",
            "affected":   [e["code"] for e in j1_entries],
            "details":    [{"code": e["code"], "from": e["from_centre"], "to": e["to_centre"]}
                           for e in j1_entries],
        })
    if j8_entries:
        assumptions_list.append({
            "question":   "J8",
            "assumption": "Cross-centre room fallback applied",
            "affected":   [e["code"] for e in j8_entries],
            "details":    [{"code": e["code"], "from": e["from_centre"], "to": e["to_centre"]}
                           for e in j8_entries],
        })

    # Assign now — after enrichment built the final lists
    stats["issues"]      = all_issues
    stats["assumptions"] = assumptions_list
    stats["true_total"]  = (stats.get("total_classes") or 0) + len(unscheduled_rooms)
    # Expose HK holidays + term starts so the UI can render the semester calendar.
    stats["hk_holidays"] = sorted(_HK_HOLIDAYS)
    stats["term_dates"]  = term_dates or {}

    del conn
    gc.collect()

    # Net shortfall warnings now generated inside assign_english_weekly() and merged above

    output_bytes = write_output_fast(
        results,
        english_weekly=english_weekly,
        issues=stats.get("issues"),
        assumptions=stats.get("assumptions"),
        term_dates=term_dates,
    )
    return output_bytes, stats


# ─── CLI entry point ──────────────────────────────────────────────────────────

def main():
    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: Input file not found:\n  {INPUT_FILE}")
        sys.exit(1)

    print(f"Input : {INPUT_FILE}")
    print(f"Output: {OUTPUT_FILE}\n")

    wb_read = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    print("[1/4] Building in-memory database...")
    conn, n_sched, n_unavail = build_db(wb_read)
    print(f"      {n_sched} classes loaded from Class list answer")
    print(f"      {n_unavail} unavailability slots loaded")

    print("[2/4] Assigning group rooms...")
    assign_group_rooms(conn)
    assigned_rooms = conn.execute(
        "SELECT COUNT(*) FROM group_rooms").fetchone()[0]
    print(f"      {assigned_rooms} class groups assigned rooms")

    print("[3/4] Assigning teachers (Lec1 / 2 / 3)...")
    unassigned, s1_warnings = assign_teachers(conn)
    results    = collect_results(conn)
    lec1_count = sum(1 for r in results if r["lec1"])
    print(f"      {lec1_count}/{len(results)} classes assigned Lec1")

    print("[4/4] Writing Excel output...")
    write_output(INPUT_FILE, OUTPUT_FILE, results)

    validate_and_report(conn, results, unassigned)


if __name__ == "__main__":
    main()
